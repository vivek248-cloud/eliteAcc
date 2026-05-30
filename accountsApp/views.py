from urllib import request
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from decimal import Decimal
from django.contrib import messages
import json
from decimal import Decimal
from django.contrib.auth.decorators import login_required

from django.db.models.functions import TruncDate
from django.utils.dateparse import parse_date

from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from zoneinfo import ZoneInfo

def login_view(request):

    # Already logged in
    if request.user.is_authenticated:
        return redirect('dashboard')

    # Show session expired message
    if request.session.pop('session_expired', False):
        messages.warning(request, "⏳ Your session expired. Please login again.")

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user and user.is_staff:
            login(request, user)

            # 🔐 Get last page safely
            redirect_to = request.session.pop('last_page', None)

            # Prevent redirect to login/logout
            blocked_paths = [
                reverse('login'),
                reverse('logout'),
            ]

            if (
                redirect_to
                and url_has_allowed_host_and_scheme(
                    redirect_to,
                    allowed_hosts={request.get_host()}
                )
                and not any(redirect_to.startswith(p) for p in blocked_paths)
            ):
                return redirect(redirect_to)

            return redirect('dashboard')

        messages.error(request, "Invalid credentials")

    return render(request, 'auth/login.html')







from django.contrib.auth import logout
from django.shortcuts import redirect

def logout_view(request):
    logout(request)
    request.session.flush()  # optional but cleaner
    return redirect('login')







from django.urls import reverse


from django.http import HttpResponseRedirect
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect

def switch_company(request, pk):

    company = get_object_or_404(Company, pk=pk)

    # ✅ Save selected company
    request.session['selected_company_id'] = company.id

    # ✅ Get previous page safely
    next_url = request.META.get('HTTP_REFERER')

    if not next_url:
        return redirect('dashboard')

    blocked_paths = [
        reverse('login'),
        reverse('logout'),
    ]

    if any(next_url.startswith(p) for p in blocked_paths):
        return redirect('dashboard')

    return HttpResponseRedirect(next_url)




# from django.utils import timezone
from zoneinfo import ZoneInfo
# from django.db.models.functions import TruncMonth
# from django.db.models import Avg


# @login_required(login_url='login')
# def home(request):

#     companies = Company.objects.all().order_by('name')



#     selected_company_id = request.session.get('selected_company_id')
#     selected_company = Company.objects.filter(
#         id=selected_company_id
#     ).first()

#     selected_bank_id = request.GET.get('bank')

#     start_date = request.GET.get('start_date')
#     end_date = request.GET.get('end_date')

#     sd = parse_date(start_date) if start_date else None
#     ed = parse_date(end_date) if end_date else None

#     if not selected_company_id:
#         return render(request, 'dashboard.html', {
#             'companies': companies,
#             'selected_company_id': None,
#         })

#     # =========================
#     # BASE QUERYSETS
#     # =========================
#     payment_qs = Payment.objects.filter(
#         client__company_id=selected_company_id
#     )

#     expense_qs = Expense.objects.filter(
#         client__company_id=selected_company_id
#     )

#     if sd:
#         payment_qs = payment_qs.filter(payment_date__gte=sd)
#         expense_qs = expense_qs.filter(expense_date__gte=sd)

#     if ed:
#         payment_qs = payment_qs.filter(payment_date__lte=ed)
#         expense_qs = expense_qs.filter(expense_date__lte=ed)

#     # =========================
#     # TOTAL PAYMENTS & EXPENSES
#     # =========================
#     totals = payment_qs.aggregate(
#         total_payments=Sum('amount')
#     )

#     total_payments = totals['total_payments'] or Decimal('0.00')

#     total_expenses = expense_qs.aggregate(
#         total=Sum('amount')
#     )['total'] or Decimal('0.00')

#     # =========================
#     # RECEIVABLES (NO LOOP)
#     # =========================
#     clients = Client.objects.filter(
#         company_id=selected_company_id
#     ).annotate(
#         paid_total=Sum('payments__amount')
#     )

#     total_balance = Decimal('0.00')

#     for c in clients:
#         paid = c.paid_total or Decimal('0.00')
#         total_balance += (c.budget - paid)

#     # =========================
#     # EXTRA DATA FOR HEALTH SYSTEM
#     # =========================
#     total_project_value = Decimal('0.00')
#     overrun_clients = 0
#     high_budget_clients = 0

#     for c in clients:
#         paid = c.paid_total or Decimal('0.00')
#         spent = c.expenses.aggregate(
#             total=Sum('amount')
#         )['total'] or Decimal('0.00')

#         total_project_value += c.budget

#         if spent > c.budget:
#             overrun_clients += 1

#         if spent > (c.budget * Decimal('0.8')):
#             high_budget_clients += 1

#     total_clients = clients.count()


#     # =========================
#     # BANKS (COMPANY SAFE)
#     # =========================
#     banks = Bank.objects.filter(
#         company_id=selected_company_id
#     )

#     for bank in banks:

#         payment_total = bank.payments.filter(
#             client__company_id=selected_company_id
#         ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

#         expense_total = bank.expenses.filter(
#             client__company_id=selected_company_id
#         ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

#         bank.available_balance = (
#             bank.opening_balance + payment_total - expense_total
#         )

#     total_bank = sum((b.available_balance for b in banks), Decimal('0.00'))

#     negative_bank_count = sum(
#     1 for b in banks if b.available_balance < 0
# )


#     # Selected bank
#     selected_bank = None
#     if selected_bank_id:
#         selected_bank = banks.filter(id=selected_bank_id).first()
#         display_bank_balance = (
#             selected_bank.available_balance if selected_bank else Decimal('0.00')
#         )
#     else:
#         display_bank_balance = total_bank



#     # =========================
#     # PROFIT
#     # =========================
#     total_profit = total_payments - total_expenses
#     profit_percentage = (
#         (total_profit / total_payments) * 100
#         if total_payments > 0 else 0
#     )




#     # ====================================
#     # 💎 FINANCIAL HEALTH SCORE (0-100)
#     # ====================================

#     # 1️⃣ Profit Score (25 max)
#     if profit_percentage > 30:
#         profit_score = 25
#     elif profit_percentage > 15:
#         profit_score = 18
#     elif profit_percentage > 5:
#         profit_score = 10
#     elif profit_percentage > 0:
#         profit_score = 5
#     else:
#         profit_score = 0

#     # 2️⃣ Cash Ratio Score
#     cash_ratio = (
#         total_bank / total_expenses
#         if total_expenses > 0 else Decimal('0')
#     )

#     if cash_ratio > 2:
#         cash_score = 25
#     elif cash_ratio > 1:
#         cash_score = 18
#     elif cash_ratio > 0.5:
#         cash_score = 10
#     else:
#         cash_score = 5

#     # 3️⃣ Receivable Ratio Score
#     receivable_ratio = (
#         total_balance / total_project_value
#         if total_project_value > 0 else Decimal('0')
#     )

#     if receivable_ratio < Decimal('0.2'):
#         receivable_score = 25
#     elif receivable_ratio < Decimal('0.4'):
#         receivable_score = 18
#     elif receivable_ratio < Decimal('0.6'):
#         receivable_score = 10
#     else:
#         receivable_score = 5

#     # 4️⃣ Budget Discipline Score
#     discipline_ratio = (
#         Decimal(overrun_clients) / Decimal(total_clients)
#         if total_clients > 0 else Decimal('0')
#     )

#     if discipline_ratio == 0:
#         discipline_score = 25
#     elif discipline_ratio < Decimal('0.1'):
#         discipline_score = 18
#     elif discipline_ratio < Decimal('0.3'):
#         discipline_score = 10
#     else:
#         discipline_score = 5

#     health_score = (
#         profit_score +
#         cash_score +
#         receivable_score +
#         discipline_score
#     )

#     if health_score >= 75:
#         health_status = "Stable"
#     elif health_score >= 50:
#         health_status = "Moderate Risk"
#     else:
#         health_status = "High Risk"







#     # ====================================
#     # 📊 TOP 5 EXPENSE CATEGORIES
#     # ====================================
#     top_categories = (
#         expense_qs
#         .values('category__name')
#         .annotate(total=Sum('amount'))
#         .order_by('-total')[:5]
#     )

#     top_category_labels = [
#         c['category__name'] or "Uncategorized"
#         for c in top_categories
#     ]

#     top_category_values = [
#         float(c['total']) for c in top_categories
#     ]



#     # ====================================
#     # 👷 SALARY DISTRIBUTION BY TEAM
#     # ====================================
#     salary_qs = expense_qs.filter(category__name__iexact='salary')

#     salary_by_team = (
#         salary_qs
#         .values('salary_to__name')
#         .annotate(total=Sum('amount'))
#         .order_by('-total')
#     )

#     salary_team_labels = [
#         s['salary_to__name'] or "Unknown"
#         for s in salary_by_team
#     ]

#     salary_team_values = [
#         float(s['total']) for s in salary_by_team
#     ]



#     # ====================================
#     # 📉 MONTHLY EXPENSE TREND
#     # ====================================
#     monthly_expenses = (
#         expense_qs
#         .annotate(month=TruncMonth('expense_date'))
#         .values('month')
#         .annotate(total=Sum('amount'))
#         .order_by('month')
#     )

#     monthly_expense_labels = [
#         m['month'].strftime("%b %Y")
#         for m in monthly_expenses
#     ]

#     monthly_expense_values = [
#         float(m['total']) for m in monthly_expenses
#     ]



#     # ====================================
#     # 🚨 SMART ALERTS (INTELLIGENT VERSION)
#     # ====================================

#     today = now().date()

#     # ---- Average Payment
#     avg_payment = payment_qs.aggregate(
#         avg=Avg('amount')
#     )['avg'] or Decimal('0.00')

#     large_payments_today = payment_qs.filter(
#         payment_date=today,
#         amount__gt=avg_payment
#     )

#     large_payment_count = large_payments_today.count()

#     # ---- Average Expense
#     avg_expense = expense_qs.aggregate(
#         avg=Avg('amount')
#     )['avg'] or Decimal('0.00')

#     high_expense_today = expense_qs.filter(
#         expense_date=today,
#         amount__gt=avg_expense
#     )

#     high_expense_count = high_expense_today.count()

#     # ---- Negative Bank Alert (already calculated earlier)
#     # negative_bank_count

#     # ---- Final Alert Count
#     alert_count = (
#         large_payment_count +
#         high_expense_count +
#         negative_bank_count
#     )



#     # =========================
#     # RECENT DATA
#     # =========================
#     recent_payments = payment_qs.select_related(
#         'client', 'bank'
#     ).order_by('-payment_date')[:5]

#     recent_expenses = expense_qs.select_related(
#         'client', 'category'
#     ).order_by('-expense_date')[:5]

#     # =========================
#     # RECENT CYCLE TOTAL
#     # =========================
#     recent_total = recent_payments.aggregate(
#         total=Sum('amount')
#     )['total'] or Decimal('0.00')

#     # =========================
#     # CURRENT MONTH RECEIVED
#     # =========================
#     today = now()

#     monthly_received_total = payment_qs.filter(
#         payment_date__year=today.year,
#         payment_date__month=today.month
#     ).aggregate(
#         total=Sum('amount')
#     )['total'] or Decimal('0.00')

#     # =========================
#     # GRAPH DATA
#     # =========================
#     payment_chart = (
#         payment_qs
#         .annotate(day=TruncDate('payment_date'))
#         .values('day')
#         .annotate(total=Sum('amount'))
#         .order_by('day')
#     )

#     expense_chart = (
#         expense_qs
#         .annotate(day=TruncDate('expense_date'))
#         .values('day')
#         .annotate(total=Sum('amount'))
#         .order_by('day')
#     )

#     payment_labels = [str(p['day']) for p in payment_chart]
#     payment_values = [float(p['total']) for p in payment_chart]

#     expense_labels = [str(e['day']) for e in expense_chart]
#     expense_values = [float(e['total']) for e in expense_chart]

#     all_dates = sorted(set(payment_labels + expense_labels))
#     profit_trend_data = []

#     for d in all_dates:
#         p = payment_values[payment_labels.index(d)] if d in payment_labels else 0
#         e = expense_values[expense_labels.index(d)] if d in expense_labels else 0
#         profit_trend_data.append(p - e)
#     current_date = now()


#     return render(request, 'dashboard.html', {
#         'companies': companies,
#         'selected_company_id': selected_company_id,

#         'clients': clients,
#         'banks': banks,

#         'total_bank': total_bank,
#         'display_bank_balance': display_bank_balance,
#         'selected_bank': selected_bank,
#         'selected_bank_id': selected_bank_id,

#         'total_balance': total_balance,
#         'total_payments': total_payments,
#         'total_expenses': total_expenses,

#         'recent_payments': recent_payments,
#         'recent_expenses': recent_expenses,

#         'start_date': start_date,
#         'end_date': end_date,

#         'total_profit': total_profit,
#         'profit_percentage': profit_percentage,
#         'recent_total': recent_total,
#         'monthly_received_total': monthly_received_total,
#         'current_date': current_date,




#         'health_score': health_score,
#         'health_status': health_status,
#         'expense_category_labels': json.dumps(top_category_labels),
#         'expense_category_values': json.dumps(top_category_values),

#         'salary_team_labels': json.dumps(salary_team_labels),
#         'salary_team_values': json.dumps(salary_team_values),

#         'monthly_expense_labels': json.dumps(monthly_expense_labels),
#         'monthly_expense_values': json.dumps(monthly_expense_values),
#         'alert_count': alert_count,

#         'large_payment_count': large_payment_count,
#         'high_expense_count': high_expense_count,
#         'avg_payment': avg_payment,
#         'avg_expense': avg_expense,
#         'negative_bank_count': negative_bank_count,




#         'profit_trend_labels': json.dumps(all_dates),
#         'profit_trend_values': json.dumps(profit_trend_data),

#         'payment_labels': json.dumps(payment_labels),
#         'payment_values': json.dumps(payment_values),
#         'expense_labels': json.dumps(expense_labels),
#         'expense_values': json.dumps(expense_values),
#     })



from django.utils import timezone
from zoneinfo import ZoneInfo
from django.db.models.functions import TruncMonth
from django.db.models import Avg
from django.utils.timezone import localtime, now
from django.db.models.functions import TruncMonth

@login_required(login_url='login')
def home(request):

    companies = Company.objects.all().order_by('name')

    selected_company_id = request.session.get('selected_company_id')
    selected_bank_id = request.GET.get('bank')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    sd = parse_date(start_date) if start_date else None
    ed = parse_date(end_date) if end_date else None

    if not selected_company_id:
        return render(request, 'dashboard.html', {
            'companies': companies,
            'selected_company_id': None,
        })

    days = request.GET.get('days')

    if days:
        try:
            days_int = int(days)
            ed = timezone.now().date()
            sd = ed - timedelta(days=days_int)
        except:
            pass

    # =========================
    # BASE QUERYSETS
    # =========================
    payment_qs = Payment.objects.filter(
        client__company_id=selected_company_id
    )

    expense_qs = Expense.objects.filter(
        client__company_id=selected_company_id
    )

    if sd:
        payment_qs = payment_qs.filter(payment_date__gte=sd)
        expense_qs = expense_qs.filter(expense_date__gte=sd)

    if ed:
        payment_qs = payment_qs.filter(payment_date__lte=ed)
        expense_qs = expense_qs.filter(expense_date__lte=ed)

    # =========================
    # TOTAL PAYMENTS / EXPENSES
    # =========================
    total_payments = payment_qs.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')

    total_expenses = expense_qs.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')

    # =========================
    # CLIENT ANNOTATIONS
    # =========================
    clients = Client.objects.filter(
        company_id=selected_company_id
    ).annotate(
        paid_total=Sum('payments__amount'),
        spent_total=Sum('expenses__amount')
    )

    total_balance = sum(
        (c.budget - (c.paid_total or Decimal('0')))
        for c in clients
    )

    total_project_value = sum(c.budget for c in clients)

    overrun_clients = sum(
        1 for c in clients
        if (c.spent_total or Decimal('0')) > c.budget
    )

    high_budget_clients = sum(
        1 for c in clients
        if (c.spent_total or Decimal('0')) > (c.budget * Decimal('0.8'))
    )

    total_clients = clients.count()

    # =========================
    # BANK BALANCES
    # =========================
    banks = Bank.objects.filter(
        company_id=selected_company_id
    ).annotate(
        payment_total=Sum('payments__amount'),
        expense_total=Sum('expenses__amount')
    )

    for bank in banks:
        payment_total = bank.payment_total or Decimal('0')
        expense_total = bank.expense_total or Decimal('0')

        bank.available_balance = (
            bank.opening_balance + payment_total - expense_total
        )

    total_bank = sum(
        (b.available_balance for b in banks),
        Decimal('0')
    )

    negative_bank_count = sum(
        1 for b in banks if b.available_balance < 0
    )

    selected_bank = None
    if selected_bank_id:
        selected_bank = banks.filter(id=selected_bank_id).first()
        display_bank_balance = (
            selected_bank.available_balance if selected_bank else Decimal('0')
        )
    else:
        display_bank_balance = total_bank

    # =========================
    # PROFIT
    # =========================
    total_profit = total_payments - total_expenses

    profit_percentage = (
        (total_profit / total_payments) * 100
        if total_payments > 0 else 0
    )

    # =========================
    # HEALTH SCORE
    # =========================
    if profit_percentage > 30:
        profit_score = 25
    elif profit_percentage > 15:
        profit_score = 18
    elif profit_percentage > 5:
        profit_score = 10
    elif profit_percentage > 0:
        profit_score = 5
    else:
        profit_score = 0

    cash_ratio = (
        total_bank / total_expenses
        if total_expenses > 0 else Decimal('0')
    )

    if cash_ratio > 2:
        cash_score = 25
    elif cash_ratio > 1:
        cash_score = 18
    elif cash_ratio > 0.5:
        cash_score = 10
    else:
        cash_score = 5

    receivable_ratio = (
        total_balance / total_project_value
        if total_project_value > 0 else Decimal('0')
    )

    if receivable_ratio < Decimal('0.2'):
        receivable_score = 25
    elif receivable_ratio < Decimal('0.4'):
        receivable_score = 18
    elif receivable_ratio < Decimal('0.6'):
        receivable_score = 10
    else:
        receivable_score = 5

    discipline_ratio = (
        Decimal(overrun_clients) / Decimal(total_clients)
        if total_clients > 0 else Decimal('0')
    )

    if discipline_ratio == 0:
        discipline_score = 25
    elif discipline_ratio < Decimal('0.1'):
        discipline_score = 18
    elif discipline_ratio < Decimal('0.3'):
        discipline_score = 10
    else:
        discipline_score = 5

    health_score = (
        profit_score +
        cash_score +
        receivable_score +
        discipline_score
    )

    if health_score >= 75:
        health_status = "Stable"
    elif health_score >= 50:
        health_status = "Moderate Risk"
    else:
        health_status = "High Risk"

    # =========================
    # TOP EXPENSE CATEGORIES
    # =========================
    top_categories = (
        expense_qs
        .values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')[:5]
    )

    top_category_labels = [
        c['category__name'] or "Uncategorized"
        for c in top_categories
    ]

    top_category_values = [
        float(c['total']) for c in top_categories
    ]

    # =========================
    # SALARY DISTRIBUTION
    # =========================
    salary_by_team = (
        expense_qs
        .filter(category__name__iexact='salary')
        .values('salary_to__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )

    salary_team_labels = [
        s['salary_to__name'] or "Unknown"
        for s in salary_by_team
    ]

    salary_team_values = [
        float(s['total']) for s in salary_by_team
    ]

    # =========================
    # MONTHLY EXPENSE TREND
    # =========================
    monthly_expenses = (
        expense_qs
        .annotate(month=TruncMonth('expense_date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    monthly_expense_labels = [
        m['month'].strftime("%b %Y")
        for m in monthly_expenses
    ]

    monthly_expense_values = [
        float(m['total']) for m in monthly_expenses
    ]

    # =========================
    # ALERTS
    # =========================
    india_tz = ZoneInfo("Asia/Kolkata")

    today = timezone.now().astimezone(india_tz).date()

    avg_payment = payment_qs.aggregate(
        avg=Avg('amount')
    )['avg'] or Decimal('0')

    avg_expense = expense_qs.aggregate(
        avg=Avg('amount')
    )['avg'] or Decimal('0')

    large_payment_count = payment_qs.filter(
        payment_date=today,
        amount__gt=avg_payment
    ).count()

    high_expense_count = expense_qs.filter(
        expense_date=today,
        amount__gt=avg_expense
    ).count()

    alert_count = (
        large_payment_count +
        high_expense_count +
        negative_bank_count
    )

    # =========================
    # RECENT DATA
    # =========================
    recent_payments = payment_qs.select_related(
        'client', 'bank'
    ).order_by('-payment_date')[:5]

    recent_expenses = expense_qs.select_related(
        'client', 'category'
    ).order_by('-expense_date')[:5]

    recent_total = recent_payments.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')

    today_dt = timezone.now().date()

    monthly_received_total = payment_qs.filter(
        payment_date__year=today_dt.year,
        payment_date__month=today_dt.month
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')

    # =========================
    # GRAPH DATA
    # =========================
    payment_chart = (
        payment_qs
        .annotate(month=TruncMonth('payment_date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    expense_chart = (
        expense_qs
        .annotate(month=TruncMonth('expense_date'))
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )

    payment_labels = [str(p['month']) for p in payment_chart]
    payment_values = [float(p['total']) for p in payment_chart]

    expense_labels = [str(e['month']) for e in expense_chart]
    expense_values = [float(e['total']) for e in expense_chart]

    all_dates = sorted(set(payment_labels + expense_labels))

    profit_trend_data = []

    for d in all_dates:
        p = payment_values[payment_labels.index(d)] if d in payment_labels else 0
        e = expense_values[expense_labels.index(d)] if d in expense_labels else 0
        profit_trend_data.append(p - e)

    current_date = timezone.now().date()

    current_hour = timezone.now().hour

    if current_hour < 12:
        greeting = "Good Morning ☀️"
    elif current_hour < 17:
        greeting = "Good Afternoon 🌤️"
    elif current_hour < 21:
        greeting = "Good Evening 🌙"
    else:
        greeting = "Working Late? 💻"

    return render(request, 'dashboard.html', {
        'companies': companies,
        'selected_company_id': selected_company_id,

        'clients': clients,
        'banks': banks,

        'total_bank': total_bank,
        'display_bank_balance': display_bank_balance,
        'selected_bank': selected_bank,
        'selected_bank_id': selected_bank_id,

        'total_balance': total_balance,
        'total_payments': total_payments,
        'total_expenses': total_expenses,

        'recent_payments': recent_payments,
        'recent_expenses': recent_expenses,

        'start_date': start_date,
        'end_date': end_date,
        'days': days,

        'total_profit': total_profit,
        'profit_percentage': profit_percentage,
        'recent_total': recent_total,
        'monthly_received_total': monthly_received_total,
        'current_date': current_date,

        'health_score': health_score,
        'health_status': health_status,

        'expense_category_labels': json.dumps(top_category_labels),
        'expense_category_values': json.dumps(top_category_values),

        'salary_team_labels': json.dumps(salary_team_labels),
        'salary_team_values': json.dumps(salary_team_values),

        'monthly_expense_labels': json.dumps(monthly_expense_labels),
        'monthly_expense_values': json.dumps(monthly_expense_values),

        'alert_count': alert_count,
        'large_payment_count': large_payment_count,
        'high_expense_count': high_expense_count,
        'avg_payment': avg_payment,
        'avg_expense': avg_expense,
        'negative_bank_count': negative_bank_count,

        'profit_trend_labels': json.dumps(all_dates),
        'profit_trend_values': json.dumps(profit_trend_data),

        'payment_labels': json.dumps(payment_labels),
        'payment_values': json.dumps(payment_values),

        'expense_labels': json.dumps(expense_labels),
        'expense_values': json.dumps(expense_values),
        'greeting': greeting,
        "today_date": timezone.localdate().strftime("%A, %d %b %Y"),  # e.g. "Monday, 30 Mar 2026"
    })




#clear alert when user visited

from django.http import JsonResponse

@login_required
def clear_alerts(request):
    request.session['alerts_seen'] = True
    return JsonResponse({"status": "ok"})



# accountsApp/views.py

from datetime import date, timedelta



def fill_date_gaps(qs):
    """
    qs: queryset with keys ['day', 'total']
    Returns: labels[], values[] with missing dates filled
    """

    data_map = {
        row['day']: float(row['total']) for row in qs
    }

    if not data_map:
        return [], []

    start_date = min(data_map.keys())
    end_date = max(data_map.keys())

    labels = []
    values = []

    current = start_date
    while current <= end_date:
        labels.append(current.strftime('%d-%m-%Y'))
        values.append(data_map.get(current, 0.0))
        current += timedelta(days=1)

    return labels, values





#index view for Company


from django.db.models import Count



from django.shortcuts import render
from django.db.models import Count, Sum
from .models import Company

def company_index(request):
    companies = Company.objects.annotate(
        total_clients=Count('clients', distinct=True),
        total_banks=Count('banks', distinct=True),
        total_worker_teams=Count('workers', distinct=True),
        total_workers=Count('workers__names', distinct=True),
    ).order_by('-id')
    
    # Calculate overall statistics
    total_companies = companies.count()
    total_all_clients = sum(c.total_clients for c in companies)
    total_all_banks = sum(c.total_banks for c in companies)
    total_all_workers = sum(c.total_workers for c in companies)
    
    context = {
        'companies': companies,
        'total_companies': total_companies,
        'total_all_clients': total_all_clients,
        'total_all_banks': total_all_banks,
        'total_all_workers': total_all_workers,
    }
    
    return render(request, 'company/index.html', context)


#create view for Company

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404

from .models import Company, ActivityLog


# =====================================================
# CREATE COMPANY
# =====================================================

def company_create(request):

    if request.method == 'POST':

        name = request.POST.get('name')

        logo = request.FILES.get('logo')

        if name:

            company = Company.objects.create(
                name=name,
                logo=logo
            )

            # =========================================
            # ACTIVITY LOG
            # =========================================

            ActivityLog.objects.create(

                action="Created Company",

                description=f"""
Company Created Successfully

Company:
{company.name}

Created By:
{request.user.username}
"""
            )

            messages.success(
                request,
                "Company created successfully"
            )

            return redirect('company_index')

        messages.error(
            request,
            "Company name is required"
        )

    return render(
        request,
        'company/create.html'
    )


# =====================================================
# DELETE COMPANY
# =====================================================

def company_delete(request, pk):

    company = get_object_or_404(
        Company,
        pk=pk
    )

    if request.method == 'POST':

        # =========================================
        # ACTIVITY LOG
        # =========================================

        ActivityLog.objects.create(

            action="Deleted Company",

            description=f"""
Company Deleted

Company:
{company.name}

Deleted By:
{request.user.username}
"""
        )

        company.delete()

        messages.success(
            request,
            "Company deleted successfully"
        )

        return redirect('company_index')

    return render(
        request,
        'company/delete.html',
        {
            'company': company
        }
    )


# =====================================================
# UPDATE COMPANY
# =====================================================

def company_update(request, pk):

    company = get_object_or_404(
        Company,
        pk=pk
    )

    if request.method == 'POST':

        old_name = company.name

        name = request.POST.get('name')

        logo = request.FILES.get('logo')

        if name:

            company.name = name

            # Only update logo if uploaded
            if logo:
                company.logo = logo

            company.save()

            # =========================================
            # ACTIVITY LOG
            # =========================================

            ActivityLog.objects.create(

                action="Updated Company",

                description=f"""
Company Updated Successfully

Old Name:
{old_name}

New Name:
{company.name}

Logo Updated:
{"Yes" if logo else "No"}

Updated By:
{request.user.username}
"""
            )

            messages.success(
                request,
                "Company updated successfully"
            )

            return redirect('company_index')

        messages.error(
            request,
            "Company name is required"
        )

    return render(
        request,
        'company/update.html',
        {
            'company': company
        }
    )



#client views will be added here

from django.db.models import Sum, Q
from decimal import Decimal
from django.contrib.auth.decorators import login_required


from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Q
from django.shortcuts import render
from decimal import Decimal


@login_required(login_url='login')
def client_index(request):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return render(request, 'client/index.html', {
            'clients': [],
            'expense_type': request.GET.get('expense_type', 'all'),
        })

    expense_type = request.GET.get('expense_type', 'all')

    clients = Client.objects.filter(
        company_id=selected_company_id
    ).select_related('company')

    client_data = []

    # =========================
    # BUILD CLIENT DATA
    # =========================
    for client in clients:

        payments_total = client.payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        expenses_qs = client.expenses.select_related('category')

        if expense_type == 'business':
            expenses_qs = expenses_qs.exclude(
                Q(category__name__iexact='home') |
                Q(category__name__iexact='personal')
            )

        elif expense_type == 'personal':
            expenses_qs = expenses_qs.filter(
                Q(category__name__iexact='home') |
                Q(category__name__iexact='personal')
            )

        expenses_total = expenses_qs.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        balance = payments_total - expenses_total
        yet_to_pay = client.budget - payments_total

        # 🔥 IMPORTANT FIX (ADD THIS)
        client_data.append({
            'id': client.id,
            'name': client.name,
            'location': client.location,
            'company': client.company,
            'budget': client.budget,
            'total_paid': payments_total,
            'total_expenses': expenses_total,
            'balance': balance,
            'yet_to_pay': yet_to_pay,
            'is_active': client.is_active,  # ✅ FIX
        })

    # =========================
    # TOTALS
    # =========================
    total_project_budget = sum(c['budget'] for c in client_data)
    total_paid = sum(c['total_paid'] for c in client_data)
    total_spend = sum(c['total_expenses'] for c in client_data)

    total_balance = total_paid - total_spend
    total_yet_to_receive = total_project_budget - total_paid

    return render(request, 'client/index.html', {
        'clients': client_data,
        'expense_type': expense_type,
        'total_project_budget': total_project_budget,
        'total_paid': total_paid,
        'total_spend': total_spend,
        'total_balance': total_balance,
        'total_yet_to_receive': total_yet_to_receive,
    })











# from decimal import Decimal, InvalidOperation
# from django.contrib.auth.decorators import login_required
# from django.contrib import messages
# from django.shortcuts import render, redirect
# from .models import Client, Company


# @login_required(login_url='login')
# def client_create(request):

#     # 🏢 COMPANY FROM SESSION
#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         messages.error(request, "Please select a company first.")
#         return redirect('dashboard')

#     company = Company.objects.filter(id=selected_company_id).first()

#     if not company:
#         messages.error(request, "Company not found.")
#         return redirect('dashboard')

#     if request.method == 'POST':
#         name = request.POST.get('name')
#         location = request.POST.get('location')
#         budget = request.POST.get('budget')

#         if not name or not budget:
#             messages.error(request, "Name and budget are required.")
#             return redirect('client_create')

#         # Normalize data (optional but recommended)
#         name = name.strip()
#         location = location.strip() if location else None

#         try:
#             budget_decimal = Decimal(budget)
#         except InvalidOperation:
#             messages.error(request, "Invalid budget amount.")
#             return redirect('client_create')

#         # 🚫 PREVENT DUPLICATE CLIENT
#         if Client.objects.filter(
#             company=company,
#             name__iexact=name,
#             location__iexact=location
#         ).exists():
#             messages.error(
#                 request,
#                 "Client with this name and location already exists."
#             )
#             return redirect('client_create')

#         # ✅ CREATE CLIENT
#         Client.objects.create(
#             company=company,
#             name=name,
#             location=location,
#             budget=budget_decimal
#         )

#         messages.success(request, "Client created successfully.")
#         return redirect('client_index')

#     return render(request, 'client/create.html', {
#         'company': company,
#     })




from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect

from .models import (
    Client,
    Company,
    ActivityLog
)


@login_required(login_url='login')
def client_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:

        messages.error(
            request,
            "Please select a company first."
        )

        return redirect('dashboard')

    company = Company.objects.filter(
        id=selected_company_id
    ).first()

    if not company:

        messages.error(
            request,
            "Company not found."
        )

        return redirect('dashboard')

    # =====================================================
    # CREATE CLIENT
    # =====================================================

    if request.method == 'POST':

        name = request.POST.get('name')

        location = request.POST.get('location')

        budget = request.POST.get('budget')

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name or not budget:

            messages.error(
                request,
                "Name and budget are required."
            )

            return redirect('client_create')

        # Clean data
        name = name.strip()

        location = (
            location.strip()
            if location else None
        )

        # =====================================================
        # BUDGET VALIDATION
        # =====================================================

        try:

            budget_decimal = Decimal(budget)

        except InvalidOperation:

            messages.error(
                request,
                "Invalid budget amount."
            )

            return redirect('client_create')

        # =====================================================
        # PREVENT DUPLICATE CLIENT
        # =====================================================

        if Client.objects.filter(
            company=company,
            name__iexact=name,
            location__iexact=location
        ).exists():

            messages.error(
                request,
                "Client with this name and location already exists."
            )

            return redirect('client_create')

        # =====================================================
        # CREATE CLIENT
        # =====================================================

        client = Client.objects.create(

            company=company,

            name=name,

            location=location,

            budget=budget_decimal
        )

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        log_activity(

    request=request,

    action="Created Client",

    action_type='Created',

    description=f"""
Client Created Successfully

Client:
{client.name}

Location:
{client.location or 'N/A'}

Budget:
Rs. {client.budget}

Company:
{company.name}
"""
)

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Client created successfully."
        )

        return redirect('client_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'client/create.html',
        {
            'company': company,
        }
    )




from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import get_object_or_404, render, redirect
from .models import Client, Company


# @login_required(login_url='login')
# def client_update(request, pk):

#     # 🏢 COMPANY FROM SESSION
#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         messages.error(request, "Please select a company first.")
#         return redirect('dashboard')

#     company = get_object_or_404(Company, id=selected_company_id)

#     # 🔒 Fetch client ONLY from selected company
#     client = get_object_or_404(
#         Client,
#         pk=pk,
#         company_id=selected_company_id
#     )

#     if request.method == 'POST':
#         name = request.POST.get('name')
#         location = request.POST.get('location')
#         budget = request.POST.get('budget')

#         if not name or not budget:
#             messages.error(request, "Name and budget are required.")
#             return redirect('client_update', pk=client.id)

#         # Normalize values
#         name = name.strip()
#         location = location.strip() if location else None

#         try:
#             budget_decimal = Decimal(budget)
#         except InvalidOperation:
#             messages.error(request, "Invalid budget amount.")
#             return redirect('client_update', pk=client.id)

#         # 🚫 Prevent duplicate client (excluding current record)
#         if Client.objects.filter(
#             company=company,
#             name__iexact=name,
#             location__iexact=location
#         ).exclude(pk=client.pk).exists():

#             messages.error(
#                 request,
#                 "Another client with this name and location already exists."
#             )
#             return redirect('client_update', pk=client.id)

#         # ✅ Update client
#         client.name = name
#         client.location = location
#         client.budget = budget_decimal
#         client.save()

#         messages.success(request, "Client updated successfully.")
#         return redirect('client_index')

#     return render(request, 'client/update.html', {
#         'client': client,
#         'company': company,
#     })



@login_required(login_url='login')
def client_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:

        messages.error(
            request,
            "Please select a company first."
        )

        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # FETCH CLIENT ONLY FROM SELECTED COMPANY
    # =====================================================

    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=selected_company_id
    )

    # =====================================================
    # UPDATE CLIENT
    # =====================================================

    if request.method == 'POST':

        # Store old values for activity log
        old_name = client.name
        old_location = client.location
        old_budget = client.budget

        name = request.POST.get('name')

        location = request.POST.get('location')

        budget = request.POST.get('budget')

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name or not budget:

            messages.error(
                request,
                "Name and budget are required."
            )

            return redirect(
                'client_update',
                pk=client.id
            )

        # Normalize values
        name = name.strip()

        location = (
            location.strip()
            if location else None
        )

        # =====================================================
        # BUDGET VALIDATION
        # =====================================================

        try:

            budget_decimal = Decimal(budget)

        except InvalidOperation:

            messages.error(
                request,
                "Invalid budget amount."
            )

            return redirect(
                'client_update',
                pk=client.id
            )

        # =====================================================
        # PREVENT DUPLICATE CLIENT
        # =====================================================

        if Client.objects.filter(
            company=company,
            name__iexact=name,
            location__iexact=location
        ).exclude(pk=client.pk).exists():

            messages.error(
                request,
                "Another client with this name and location already exists."
            )

            return redirect(
                'client_update',
                pk=client.id
            )

        # =====================================================
        # UPDATE CLIENT
        # =====================================================

        client.name = name
        client.location = location
        client.budget = budget_decimal

        client.save()

        # =====================================================
        # ACTIVITY LOG
        # =====================================================
        log_activity(

            request=request,

            action="Updated Client",

            action_type='Updated',

            description=f"""
        Client Updated Successfully

        Old Name:
        {old_name}

        New Name:
        {client.name}

        Old Location:
        {old_location or 'N/A'}

        New Location:
        {client.location or 'N/A'}

        Old Budget:
        Rs. {old_budget}

        New Budget:
        Rs. {client.budget}

        Company:
        {company.name}
        """
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Client updated successfully."
        )

        return redirect('client_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'client/update.html',
        {
            'client': client,
            'company': company,
        }
    )






@login_required(login_url='login')
def client_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:

        messages.error(
            request,
            "Please select a company first."
        )

        return redirect('dashboard')

    # =====================================================
    # FETCH CLIENT ONLY FROM SELECTED COMPANY
    # =====================================================

    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=selected_company_id
    )

    # =====================================================
    # DELETE CLIENT
    # =====================================================

    if request.method == 'POST':

        # =========================================
        # ACTIVITY LOG
        # =========================================

        log_activity(

    request=request,

    action="Deleted Client",

    action_type='Deleted',

    description=f"""
Client Deleted Successfully

Client:
{client.name}

Location:
{client.location or 'N/A'}

Budget:
Rs. {client.budget}

Company:
{client.company.name}
"""
)
        # =========================================
        # DELETE CLIENT
        # =========================================

        client.delete()

        # =========================================
        # SUCCESS MESSAGE
        # =========================================

        messages.success(
            request,
            "Client deleted successfully."
        )

        return redirect('client_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'client/delete.html',
        {
            'client': client
        }
    )



from django.utils.dateparse import parse_date

from decimal import Decimal


from decimal import Decimal
from django.shortcuts import get_object_or_404, render, redirect
from django.utils.dateparse import parse_date


def client_info(request, pk):

    # =========================
    # 🏢 COMPANY SESSION CHECK
    # =========================
    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    # 🔐 Client must belong to selected company
    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=selected_company_id
    )

    payments_qs = client.payments.select_related('bank')

    expenses_qs = client.expenses.select_related(
        'bank',
        'category',
        'subcategory',
        'salary_to',
        'worker_name'
    )

    # =========================
    # 🔍 FILTER VALUES
    # =========================
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    spend_mode = request.GET.get('spend_mode')
    category_id = request.GET.get('category')
    subcategory_id = request.GET.get('subcategory')
    worker_id = request.GET.get('worker')
    worker_name_id = request.GET.get('worker_name')

    # =========================
    # 📅 DATE FILTERS
    # =========================
    if start_date:
        sd = parse_date(start_date)
        expenses_qs = expenses_qs.filter(expense_date__gte=sd)

    if end_date:
        ed = parse_date(end_date)
        expenses_qs = expenses_qs.filter(expense_date__lte=ed)

    # =========================
    # 💳 SPEND MODE
    # =========================
    if spend_mode:
        expenses_qs = expenses_qs.filter(spend_mode=spend_mode)

    # =========================
    # 🏷 CATEGORY (company safe)
    # =========================
    if category_id:
        expenses_qs = expenses_qs.filter(
            category_id=category_id,
            category__company_id=selected_company_id
        )

    # =========================
    # 🏷 SUBCATEGORY (company safe)
    # =========================
    if subcategory_id:
        expenses_qs = expenses_qs.filter(
            subcategory_id=subcategory_id,
            subcategory__company_id=selected_company_id
        )

    # =========================
    # 👷 WORKER TEAM
    # =========================
    if worker_id:
        expenses_qs = expenses_qs.filter(
            salary_to_id=worker_id,
            salary_to__company_id=selected_company_id
        )

    # =========================
    # 👤 WORKER NAME
    # =========================
    if worker_name_id:
        expenses_qs = expenses_qs.filter(
            worker_name_id=worker_name_id,
            worker_name__worker__company_id=selected_company_id
        )

    # =========================
    # ORDER
    # =========================
    payments_qs = payments_qs.order_by('payment_date', 'id')
    expenses_qs = expenses_qs.order_by('expense_date', 'id')

    # =========================
    # 💰 PAYMENT CUMULATIVE
    # =========================
    running_paid = Decimal('0.00')
    payment_rows = []

    for p in payments_qs:
        before = running_paid
        running_paid += p.amount
        remaining = client.budget - running_paid

        payment_rows.append({
            'date': p.payment_date,
            'amount': p.amount,
            'mode': p.payment_mode,
            'bank': p.bank.name if p.bank else 'Cash',
            'before': before,
            'remaining': remaining,
        })

    # =========================
    # 🔴 EXPENSE CUMULATIVE
    # =========================
    total_paid = running_paid
    running_spent = Decimal('0.00')
    expense_rows = []

    for e in expenses_qs:
        before = running_spent
        running_spent += e.amount
        remaining = total_paid - running_spent

        if e.salary_to:
            team_name = e.salary_to.name
            worker_name = e.worker_name.name if e.worker_name else None
            worker_display = f"{team_name} / {worker_name}" if worker_name else team_name
        else:
            worker_display = None

        expense_rows.append({
            'date': e.expense_date,
            'description': e.description,
            'category': e.category.name if e.category else None,
            'subcategory': e.subcategory.name if e.subcategory else None,
            'before': before,
            'now': e.amount,
            'remaining': remaining,
            'mode': e.spend_mode,
            'bank': e.bank.name if e.bank else 'Cash',
            'worker': worker_display,
        })

    # ✅ Correct totals AFTER loop
    total_spend = running_spent
    net_balance = total_paid - total_spend
    remaining_amount = client.budget - total_paid

    return render(request, 'client/clientinfo.html', {
        'client': client,
        'payments': payment_rows,
        'expenses': expense_rows,

        # filters
        'start_date': start_date,
        'end_date': end_date,
        'spend_mode': spend_mode,
        'total_spend': total_spend,
        'net_balance': net_balance,
        'remaining_amount': remaining_amount,
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
        'selected_worker': worker_id,
        'selected_worker_name': worker_name_id,

        # dropdown data
        'categories': ExpenseCategory.objects.filter(
            company_id=selected_company_id
        ),

        'subcategories': ExpenseSubCategory.objects.filter(
            company_id=selected_company_id
        ).select_related('category'),

        'workers': Worker.objects.filter(
            company_id=selected_company_id
        ),

        'worker_names': WorkerName.objects.filter(
            worker__company_id=selected_company_id
        ).select_related('worker'),
    })





# export pdf views will be added here



from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.http import HttpResponse
from decimal import Decimal
from django.utils.dateparse import parse_date
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Sum


def client_info_pdf(request, pk):

    # =========================
    # 🏢 COMPANY SESSION CHECK
    # =========================
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=selected_company_id
    )

    payments_qs = client.payments.select_related('bank')
    expenses_qs = client.expenses.select_related(
        'bank', 'salary_to', 'worker_name'
    )

    # =========================
    # 🔍 FILTERS
    # =========================
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    payment_mode = request.GET.get('payment_mode')
    spend_mode = request.GET.get('spend_mode')

    if start_date:
        sd = parse_date(start_date)
        payments_qs = payments_qs.filter(payment_date__gte=sd)
        expenses_qs = expenses_qs.filter(expense_date__gte=sd)

    if end_date:
        ed = parse_date(end_date)
        payments_qs = payments_qs.filter(payment_date__lte=ed)
        expenses_qs = expenses_qs.filter(expense_date__lte=ed)

    if payment_mode:
        payments_qs = payments_qs.filter(payment_mode=payment_mode)

    if spend_mode:
        expenses_qs = expenses_qs.filter(spend_mode=spend_mode)

    payments_qs = payments_qs.order_by('payment_date', 'id')
    expenses_qs = expenses_qs.order_by('expense_date', 'id')

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        'header_style',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=10
    )

    # =========================
    # 💰 PAYMENTS
    # =========================
    running_paid = Decimal('0.00')
    payment_rows = []

    for p in payments_qs:
        before = running_paid
        running_paid += p.amount
        remaining = client.budget - running_paid

        payment_rows.append([
            p.payment_date.strftime('%d-%m-%Y'),
            f"Rs. {before:,.2f}",
            f"Rs. {p.amount:,.2f}",
            f"Rs. {remaining:,.2f}",
            p.payment_mode.capitalize(),
            p.bank.name if p.bank else 'Cash'
        ])

    total_paid = running_paid

    # =========================
    # 🔴 EXPENSES
    # =========================
    running_spent = Decimal('0.00')
    expense_rows = []

    for e in expenses_qs:
        before = running_spent
        running_spent += e.amount
        remaining = total_paid - running_spent

        # 👷 Worker Display
        if e.salary_to:
            team = e.salary_to.name
            name = e.worker_name.name if e.worker_name else None
            worker_display = f"{team} / {name}" if name else team
        else:
            worker_display = '—'

        expense_rows.append([
            e.expense_date.strftime('%d-%m-%Y'),
            e.category.name if e.category else '—',
            e.description,
            worker_display,
            f"Rs.{e.amount:,.2f}",
            f"Rs.{remaining:,.2f}",
            e.spend_mode.capitalize(),
            e.bank.name if e.bank else 'Cash'
        ])

    total_spent = running_spent
    final_balance = total_paid - total_spent

    # =========================
    # 📄 PDF SETUP (LANDSCAPE)
    # =========================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{client.name}_statement.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )

    elements = []

    # =========================
    # 🧾 HEADER
    # =========================
    elements.append(Paragraph("CLIENT FINANCIAL STATEMENT", header_style))
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(f"<b>Client:</b> {client.name}", styles['Normal'])
    )
    elements.append(
        Paragraph(f"<b>Project Value:</b> Rs. {client.budget:,.2f}", styles['Normal'])
    )
    elements.append(Spacer(1, 20))

    # =========================
    # 💰 PAYMENTS TABLE
    # =========================
    elements.append(Paragraph("PAYMENTS", styles['Heading2']))
    elements.append(Spacer(1, 8))

    payment_table = Table(
        [['Date', 'Before', 'Paid', 'Remaining', 'Mode', 'Bank']]
        + payment_rows,
        colWidths=[90, 110, 110, 120, 90, 120]
    )

    payment_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.whitesmoke, colors.transparent]),
    ]))

    elements.append(payment_table)
    elements.append(Spacer(1, 20))

    # =========================
    # 🔴 EXPENSES TABLE
    # =========================
    elements.append(Paragraph("EXPENSES", styles['Heading2']))
    elements.append(Spacer(1, 8))

    expense_table = Table(
        [['Date', 'Category','Description', 'Worker', 'Spent', 'Remaining', 'Mode', 'Bank']]
        + expense_rows,
        colWidths=[70, 100, 140, 110, 95, 110, 70, 90]
    )

    expense_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7A1F1F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.whitesmoke, colors.transparent]),
    ]))

    elements.append(expense_table)
    elements.append(Spacer(1, 25))

    # =========================
    # 📦 SUMMARY BOX
    # =========================
    summary_table = Table(
        [
            ['TOTAL PAID', f"Rs. {total_paid:,.2f}"],
            ['TOTAL SPENT', f"Rs. {total_spent:,.2f}"],
            ['FINAL BALANCE', f"Rs. {final_balance:,.2f}"],
        ],
        colWidths=[250, 200]
    )

    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F2F2F2')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('PADDING', (0, 0), (-1, -1), 10),
    ]))

    elements.append(summary_table)

    doc.build(elements)
    return response


#export as excel will be added here

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def client_info_excel(request, pk):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=selected_company_id
    )

    payments_qs = client.payments.select_related('bank')
    expenses_qs = client.expenses.select_related(
        'bank',
        'salary_to',
        'worker_name'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Statement"

    # =========================
    # TITLE
    # =========================

    ws.merge_cells('A1:H1')

    title_cell = ws['A1']
    title_cell.value = "CLIENT FINANCIAL STATEMENT"

    title_cell.font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    title_cell.fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    title_cell.alignment = Alignment(horizontal="center")

    # =========================
    # CLIENT INFO
    # =========================

    ws['A3'] = "Client Name"
    ws['B3'] = client.name

    ws['A4'] = "Project Value"
    ws['B4'] = float(client.budget)

    # =========================
    # PAYMENTS HEADER
    # =========================

    payment_start = 7

    headers = [
        'Date',
        'Before',
        'Paid',
        'Remaining',
        'Mode',
        'Bank'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(payment_start, col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )

    running_paid = 0

    row = payment_start + 1

    for p in payments_qs:

        before = running_paid
        running_paid += p.amount
        remaining = client.budget - running_paid

        ws.cell(row, 1, p.payment_date.strftime('%d-%m-%Y'))
        ws.cell(row, 2, float(before))
        ws.cell(row, 3, float(p.amount))
        ws.cell(row, 4, float(remaining))
        ws.cell(row, 5, p.payment_mode.capitalize())
        ws.cell(row, 6, p.bank.name if p.bank else "Cash")

        row += 1

    total_paid = running_paid

    # =========================
    # EXPENSES
    # =========================

    row += 3

    expense_header_row = row

    expense_headers = [
        'Date',
        'Category',
        'Description',
        'Worker',
        'Spent',
        'Remaining',
        'Mode',
        'Bank'
    ]

    for col_num, header in enumerate(expense_headers, 1):
        cell = ws.cell(expense_header_row, col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="C0504D",
            end_color="C0504D",
            fill_type="solid"
        )

    running_spent = 0

    row += 1

    for e in expenses_qs:

        before = running_spent
        running_spent += e.amount
        remaining = total_paid - running_spent

        if e.salary_to:
            team = e.salary_to.name
            name = e.worker_name.name if e.worker_name else None
            worker_display = f"{team} / {name}" if name else team
        else:
            worker_display = '—'

        ws.cell(row, 1, e.expense_date.strftime('%d-%m-%Y'))
        ws.cell(row, 2, e.category.name if e.category else '—')
        ws.cell(row, 3, e.description)
        ws.cell(row, 4, worker_display)
        ws.cell(row, 5, float(e.amount))
        ws.cell(row, 6, float(remaining))
        ws.cell(row, 7, e.spend_mode.capitalize())
        ws.cell(row, 8, e.bank.name if e.bank else "Cash")

        row += 1

    total_spent = running_spent
    final_balance = total_paid - total_spent

    # =========================
    # SUMMARY
    # =========================

    row += 2

    ws.cell(row, 1, "TOTAL PAID")
    ws.cell(row, 2, float(total_paid))

    row += 1

    ws.cell(row, 1, "TOTAL SPENT")
    ws.cell(row, 2, float(total_spent))

    row += 1

    ws.cell(row, 1, "FINAL BALANCE")
    ws.cell(row, 2, float(final_balance))

    # =========================
    # COLUMN WIDTHS
    # =========================

    widths = {
        1: 18,
        2: 20,
        3: 28,
        4: 30,
        5: 18,
        6: 20,
        7: 15,
        8: 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{client.name}_statement.xlsx"'
    )

    wb.save(response)

    return response






from decimal import Decimal
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.db.models import Sum
from .models import Client, Payment, Expense


from decimal import Decimal
from itertools import chain
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.db.models import Sum, F, Value
from django.db.models.functions import Coalesce
from .models import Client, Payment, Expense


from decimal import Decimal
from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date
from django.contrib.auth.decorators import login_required


@login_required(login_url='login')
def all_client_index(request):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    order = request.GET.get('order', 'new')
    txn_type = request.GET.get('txn_type', 'all')

    sd = parse_date(start_date) if start_date else None
    ed = parse_date(end_date) if end_date else None

    rows = []

    clients = Client.objects.filter(
        company_id=selected_company_id
    ).select_related('company')

    # =========================
    # BUILD ROWS
    # =========================
    for client in clients:

        payments = client.payments.all()
        expenses = client.expenses.all()

        if sd:
            payments = payments.filter(payment_date__gte=sd)
            expenses = expenses.filter(expense_date__gte=sd)

        if ed:
            payments = payments.filter(payment_date__lte=ed)
            expenses = expenses.filter(expense_date__lte=ed)

        payments = payments.order_by('payment_date', 'id')
        expenses = expenses.order_by('expense_date', 'id')

        running_paid = Decimal('0.00')
        running_spent = Decimal('0.00')

        # PAYMENTS
        for p in payments:
            before_paid = running_paid
            running_paid += p.amount

            rows.append({
                'date': p.payment_date,
                'client': client.name,
                'company': client.company.name,
                'budget': client.budget,
                'previous_paid': before_paid,
                'paid_now': p.amount,
                'yet_to_pay': client.budget - running_paid,
                'total_paid': running_paid,
                'spend_detail': '—',
                'salary_info': '—',
                'spend_amount': Decimal('0.00'),
                'balance': running_paid - running_spent,
                'type': 'payment',
            })

        # EXPENSES
        for e in expenses:
            running_spent += e.amount

            if e.category and e.category.name.lower() == 'salary':
                team_name = e.salary_to.name if e.salary_to else ''
                worker_name = e.worker_name.name if e.worker_name else ''
                salary_info = f"{team_name}"
                if worker_name:
                    salary_info += f" / {worker_name}"
            else:
                salary_info = '—'

            rows.append({
                'date': e.expense_date,
                'client': client.name,
                'company': client.company.name,
                'budget': client.budget,
                'previous_paid': running_paid,
                'paid_now': Decimal('0.00'),
                'yet_to_pay': client.budget - running_paid,
                'total_paid': running_paid,
                'spend_detail': e.description,
                'salary_info': salary_info,
                'spend_amount': e.amount,
                'balance': running_paid - running_spent,
                'type': 'expense',
            })

    # =========================
    # TOTAL CALCULATIONS (CORRECT LOCATION)
    # =========================

    total_project_value = Decimal('0.00')
    total_paid = Decimal('0.00')
    total_spend = Decimal('0.00')
    total_yet_to_pay = Decimal('0.00')

    for client in clients:

        client_payments = client.payments.all()
        client_expenses = client.expenses.all()

        if sd:
            client_payments = client_payments.filter(payment_date__gte=sd)
            client_expenses = client_expenses.filter(expense_date__gte=sd)

        if ed:
            client_payments = client_payments.filter(payment_date__lte=ed)
            client_expenses = client_expenses.filter(expense_date__lte=ed)

        client_paid_total = client_payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        client_spend_total = client_expenses.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        total_project_value += client.budget
        total_paid += client_paid_total
        total_spend += client_spend_total
        total_yet_to_pay += (client.budget - client_paid_total)

    grand_balance = total_paid - total_spend

    # =========================
    # FILTER TRANSACTION TYPE
    # =========================
    if txn_type != 'all':
        rows = [r for r in rows if r['type'] == txn_type]

    # SORT
    rows = sorted(
        rows,
        key=lambda x: x['date'],
        reverse=(order == 'new')
    )

    return render(request, 'client/all_clients_statement.html', {
        'rows': rows,
        'start_date': start_date,
        'end_date': end_date,
        'order': order,
        'txn_type': txn_type,
        'total_project_value': total_project_value,
        'total_paid': total_paid,
        'total_spend': total_spend,
        'total_yet_to_pay': total_yet_to_pay,
        'grand_balance': grand_balance,
    })









from decimal import Decimal
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors


def clean_date(value):
    if not value or value == 'None':
        return None
    return parse_date(value)


from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import PageBreak

def add_page_numbers(canvas, doc):
    canvas.drawRightString(
        820, 15,
        f"Page {doc.page}"
    )



from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import redirect
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def all_client_info_pdf(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    start_date = clean_date(request.GET.get('start_date'))
    end_date = clean_date(request.GET.get('end_date'))
    order = request.GET.get('order', 'new')
    txn_type = request.GET.get('txn_type', 'all')

    clients = Client.objects.filter(
        company_id=selected_company_id
    ).select_related('company')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="all_clients_statement.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=15,
        leftMargin=15,
        topMargin=20,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    # 🔥 SMALL TEXT STYLE
    small = ParagraphStyle(
        name='Small',
        fontSize=8,
        leading=10
    )

    bold = ParagraphStyle(
        name='Bold',
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold'
    )

    elements = []

    elements.append(Paragraph("<b>All Clients Financial Statement</b>", styles['Title']))
    elements.append(Spacer(1, 10))

    # =============================
    # TABLE DATA
    # =============================
    table_data = [[
        Paragraph('Date', bold),
        Paragraph('Client', bold),
        Paragraph('Company', bold),
        Paragraph('Prev Paid', bold),
        Paragraph('Paid Now', bold),
        Paragraph('Yet To Pay', bold),
        Paragraph('Spend Detail', bold),
        Paragraph('Spend Amount', bold),
        Paragraph('Balance', bold),
    ]]

    row_styles = []
    row_index = 1

    grand_paid = Decimal('0.00')
    grand_spent = Decimal('0.00')

    for client in clients:

        payments = client.payments.all()
        expenses = client.expenses.all()

        if start_date:
            payments = payments.filter(payment_date__gte=start_date)
            expenses = expenses.filter(expense_date__gte=start_date)

        if end_date:
            payments = payments.filter(payment_date__lte=end_date)
            expenses = expenses.filter(expense_date__lte=end_date)

        ledger = []

        for p in payments:
            ledger.append({'type': 'payment', 'date': p.payment_date, 'amount': p.amount, 'obj': p})

        for e in expenses:
            ledger.append({'type': 'expense', 'date': e.expense_date, 'amount': e.amount, 'obj': e})

        ledger = sorted(ledger, key=lambda x: x['date'], reverse=(order == 'new'))

        running_paid = Decimal('0.00')
        running_spent = Decimal('0.00')

        for entry in ledger:

            if txn_type != 'all' and entry['type'] != txn_type:
                continue

            prev_paid = running_paid

            if entry['type'] == 'payment':

                running_paid += entry['amount']
                grand_paid += entry['amount']

                yet_to_pay = client.budget - running_paid
                balance = running_paid - running_spent

                table_data.append([
                    Paragraph(entry['date'].strftime('%d-%m-%Y'), small),
                    Paragraph(client.name, small),
                    Paragraph(client.company.name, small),
                    Paragraph(f"Rs. {prev_paid:.2f}", small),
                    Paragraph(f"Rs. {entry['amount']:.2f}", small),
                    Paragraph(f"Rs. {yet_to_pay:.2f}", small),
                    Paragraph('—', small),
                    Paragraph('—', small),
                    Paragraph(f"Rs. {balance:.2f}", small),
                ])

            else:

                running_spent += entry['amount']
                grand_spent += entry['amount']

                yet_to_pay = client.budget - running_paid
                balance = running_paid - running_spent

                table_data.append([
                    Paragraph(entry['date'].strftime('%d-%m-%Y'), small),
                    Paragraph(client.name, small),
                    Paragraph(client.company.name, small),
                    Paragraph(f"Rs. {prev_paid:.2f}", small),
                    Paragraph('—', small),
                    Paragraph(f"Rs. {yet_to_pay:.2f}", small),
                    Paragraph(entry['obj'].description or '—', small),
                    Paragraph(f"Rs. {entry['amount']:.2f}", small),
                    Paragraph(f"Rs. {balance:.2f}", small),
                ])

            # 🔴 NEGATIVE COLORS
            if yet_to_pay < 0:
                row_styles.append(('TEXTCOLOR', (5, row_index), (5, row_index), colors.red))

            if balance < 0:
                row_styles.append(('TEXTCOLOR', (8, row_index), (8, row_index), colors.red))

            row_index += 1

        # =========================
        # CLIENT TOTAL
        # =========================
        client_balance = running_paid - running_spent

        table_data.append([
            Paragraph(f"{client.name} TOTAL", bold),
            '', '', '',
            Paragraph(f"Rs. {running_paid:.2f}", bold),
            '',
            '',
            Paragraph(f"Rs. {running_spent:.2f}", bold),
            Paragraph(f"Rs. {client_balance:.2f}", bold),
        ])

        row_styles.append(('FONTNAME', (0, row_index), (-1, row_index), 'Helvetica-Bold'))

        if client_balance < 0:
            row_styles.append(('TEXTCOLOR', (8, row_index), (8, row_index), colors.red))

        row_index += 1

    # =========================
    # GRAND TOTAL
    # =========================
    grand_balance = grand_paid - grand_spent

    table_data.append([
        Paragraph('GRAND TOTAL', bold),
        '', '', '',
        Paragraph(f"Rs. {grand_paid:.2f}", bold),
        '',
        '',
        Paragraph(f"Rs. {grand_spent:.2f}", bold),
        Paragraph(f"Rs. {grand_balance:.2f}", bold),
    ])

    row_styles.append(('BACKGROUND', (0, row_index), (-1, row_index), colors.lightgrey))
    row_styles.append(('FONTNAME', (0, row_index), (-1, row_index), 'Helvetica-Bold'))

    if grand_balance < 0:
        row_styles.append(('TEXTCOLOR', (8, row_index), (8, row_index), colors.red))

    # =========================
    # TABLE
    # =========================
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[65,80, 100, 70, 70, 75, 160, 75, 75]
    )

    table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (3,1), (-1,-1), 'RIGHT'),

        # padding fix
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('BOTTOMPADDING', (0,0), (-1,-1), 3),
    ] + row_styles))

    elements.append(table)

    doc.build(
        elements,
        onFirstPage=add_page_numbers,
        onLaterPages=add_page_numbers
    )

    return response




#export as excel for all clients will be added here

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter

def all_client_info_excel(request):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    start_date = clean_date(request.GET.get('start_date'))
    end_date = clean_date(request.GET.get('end_date'))
    order = request.GET.get('order', 'new')
    txn_type = request.GET.get('txn_type', 'all')

    clients = Client.objects.filter(
        company_id=selected_company_id
    ).select_related('company')

    wb = Workbook()
    ws = wb.active
    ws.title = "All Clients Statement"

    # =========================
    # STYLES
    # =========================

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    total_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    red_font = Font(
        color="FF0000",
        bold=True
    )

    white_bold = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(bold=True)

    thin = Side(style='thin', color='CCCCCC')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    # =========================
    # TITLE
    # =========================

    ws.merge_cells('A1:J1')

    title = ws['A1']

    title.value = "ALL CLIENTS FINANCIAL STATEMENT"

    title.font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    title.fill = header_fill
    title.alignment = center

    # =========================
    # HEADERS
    # =========================

    headers = [
        'Date',
        'Client',
        'Location',
        'Company',
        'Prev Paid',
        'Paid Now',
        'Yet To Pay',
        'Spend Detail',
        'Spend Amount',
        'Balance',
    ]

    row = 3

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row, col_num)

        cell.value = header
        cell.font = white_bold
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row += 1

    grand_paid = Decimal('0.00')
    grand_spent = Decimal('0.00')

    # =========================
    # DATA
    # =========================

    for client in clients:

        payments = client.payments.all()
        expenses = client.expenses.all()

        if start_date:
            payments = payments.filter(payment_date__gte=start_date)
            expenses = expenses.filter(expense_date__gte=start_date)

        if end_date:
            payments = payments.filter(payment_date__lte=end_date)
            expenses = expenses.filter(expense_date__lte=end_date)

        ledger = []

        for p in payments:
            ledger.append({
                'type': 'payment',
                'date': p.payment_date,
                'amount': p.amount,
                'obj': p
            })

        for e in expenses:
            ledger.append({
                'type': 'expense',
                'date': e.expense_date,
                'amount': e.amount,
                'obj': e
            })

        ledger = sorted(
            ledger,
            key=lambda x: x['date'],
            reverse=(order == 'new')
        )

        running_paid = Decimal('0.00')
        running_spent = Decimal('0.00')

        for entry in ledger:

            if txn_type != 'all' and entry['type'] != txn_type:
                continue

            prev_paid = running_paid

            if entry['type'] == 'payment':

                running_paid += entry['amount']
                grand_paid += entry['amount']

                yet_to_pay = client.budget - running_paid
                balance = running_paid - running_spent

                data = [
                    entry['date'].strftime('%d-%m-%Y'),
                    client.name,
                    client.location or '—',
                    client.company.name,
                    float(prev_paid),
                    float(entry['amount']),
                    float(yet_to_pay),
                    '—',
                    '—',
                    float(balance)
                ]

            else:

                running_spent += entry['amount']
                grand_spent += entry['amount']

                yet_to_pay = client.budget - running_paid
                balance = running_paid - running_spent

                data = [
                    entry['date'].strftime('%d-%m-%Y'),
                    client.name,
                    client.location or '—',
                    client.company.name,
                    float(prev_paid),
                    '—',
                    float(yet_to_pay),
                    entry['obj'].description or '—',
                    float(entry['amount']),
                    float(balance)
                ]

            for col_num, value in enumerate(data, 1):

                cell = ws.cell(row, col_num)

                cell.value = value
                cell.border = border

                if col_num >= 5:
                    cell.alignment = right

            # Negative values
            if yet_to_pay < 0:
                ws.cell(row, 7).font = red_font

            if balance < 0:
                ws.cell(row, 10).font = red_font

            row += 1

        # =========================
        # CLIENT TOTAL ROW
        # =========================

        client_balance = running_paid - running_spent

        totals = [
            f"{client.name} TOTAL",
            '',
            '',
            '',
            float(running_paid),
            '',
            '',
            '',
            float(running_spent),
            float(client_balance)
        ]

        for col_num, value in enumerate(totals, 1):

            cell = ws.cell(row, col_num)

            cell.value = value
            cell.font = bold_font
            cell.fill = total_fill
            cell.border = border

            if col_num >= 5:
                cell.alignment = right

        if client_balance < 0:
            ws.cell(row, 10).font = red_font

        row += 1

    # =========================
    # GRAND TOTAL ROW
    # =========================

    grand_balance = grand_paid - grand_spent

    grand = [
        'GRAND TOTAL',
        '',
        '',
        '',
        float(grand_paid),
        '',
        '',
        '',
        float(grand_spent),
        float(grand_balance)
    ]

    for col_num, value in enumerate(grand, 1):

        cell = ws.cell(row, col_num)

        cell.value = value

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = header_fill
        cell.border = border

        if col_num >= 5:
            cell.alignment = right

    if grand_balance < 0:
        ws.cell(row, 10).font = Font(
            bold=True,
            color="FF0000"
        )

    # =========================
    # COLUMN WIDTHS
    # =========================

    widths = {
        1: 15,   # Date
        2: 28,   # Client
        3: 28,   # Location
        4: 32,   # Company
        5: 18,   # Prev Paid
        6: 18,   # Paid Now
        7: 18,   # Yet To Pay
        8: 40,   # Spend Detail
        9: 18,   # Spend Amount
        10: 18,  # Balance
    }

    for col, width in widths.items():
        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename="all_clients_statement.xlsx"'
    )

    wb.save(response)

    return response





#bank views will be added here


from django.db.models import Sum, Max, Q
from django.utils import timezone
from zoneinfo import ZoneInfo
from decimal import Decimal
from django.shortcuts import redirect


def bank_index(request):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    # =========================
    # 🔍 FIND BANKS USED BY THIS COMPANY
    # =========================

    banks = Bank.objects.filter(
        company_id=selected_company_id
    )

    # =========================
    # 🔘 FILTERS
    # =========================
    selected_bank = request.GET.get('bank')
    filter_type = request.GET.get('filter_type')
    filter_date = request.GET.get('date')
    filter_month = request.GET.get('month')
    filter_year = request.GET.get('year')

    if selected_bank:
        banks = banks.filter(id=selected_bank)

    total_bank = Decimal('0.00')

    for bank in banks:

        # =========================
        # 📊 BASE QUERYSETS
        # =========================
        payments = bank.payments.filter(
            client__company_id=selected_company_id
        )

        expenses = bank.expenses.filter(
            client__company_id=selected_company_id
        )

        # =========================
        # 📅 DATE FILTERS
        # =========================
        if filter_type == 'day' and filter_date:
            payments = payments.filter(payment_date=filter_date)
            expenses = expenses.filter(expense_date=filter_date)

        elif filter_type == 'month' and filter_month and filter_year:
            payments = payments.filter(
                payment_date__month=int(filter_month),
                payment_date__year=int(filter_year)
            )
            expenses = expenses.filter(
                expense_date__month=int(filter_month),
                expense_date__year=int(filter_year)
            )

        elif filter_type == 'year' and filter_year:
            payments = payments.filter(payment_date__year=int(filter_year))
            expenses = expenses.filter(expense_date__year=int(filter_year))

        # =========================
        # 🔥 COUNT (ASSIGN TO BANK)
        # =========================
        bank.payment_count = payments.count()
        bank.expense_count = expenses.count()
        bank.transaction_count = bank.payment_count + bank.expense_count

        # =========================
        # 💰 TOTALS
        # =========================
        bank.payment_total = payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        bank.expense_total = expenses.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        # =========================
        # 💰 BALANCE
        # =========================
        bank.filtered_balance = bank.available_balance

        # =========================
        # 🗓 LAST TRANSACTION DATE
        # =========================
        last_payment_date = payments.aggregate(
            d=Max('payment_date')
        )['d']

        last_expense_date = expenses.aggregate(
            d=Max('expense_date')
        )['d']

        bank.last_transaction_date = max(
            filter(None, [last_payment_date, last_expense_date]),
            default=None
        )

        # =========================
        # ➕ TOTAL BANK
        # =========================
        total_bank += bank.filtered_balance


    return render(request, 'bank/index.html', {
        'banks': banks,
        'total_bank': total_bank,

        'selected_bank': selected_bank,
        'filter_type': filter_type,
        'filter_date': filter_date,
        'filter_month': filter_month,
        'filter_year': filter_year,

        'months': range(1, 13),
        'current_year': timezone.now().year,
    })







from decimal import Decimal
from django.shortcuts import render, redirect
from django.contrib import messages


# def bank_create(request):

#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         return redirect('dashboard')

#     if request.method == 'POST':

#         name = request.POST.get('name', '').strip()
#         opening_balance = Decimal(request.POST.get('opening_balance') or 0)

#         if not name:
#             messages.error(request, "Bank name is required.")
#             return redirect('bank_create')

#         # 🔐 Prevent duplicate bank inside same company
#         if Bank.objects.filter(
#             company_id=selected_company_id,
#             name__iexact=name,
#             is_active=True
#         ).exists():
#             messages.error(request, "Bank with this name already exists.")
#             return redirect('bank_create')

#         Bank.objects.create(
#             company_id=selected_company_id,
#             name=name,
#             opening_balance=opening_balance,
#             available_balance=opening_balance
#         )

#         messages.success(request, "Bank created successfully.")
#         return redirect('bank_index')

#     return render(request, 'bank/create.html')



@login_required(login_url='login')
def bank_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # CREATE BANK
    # =====================================================

    if request.method == 'POST':

        name = request.POST.get(
            'name',
            ''
        ).strip()

        opening_balance = Decimal(
            request.POST.get('opening_balance') or 0
        )

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Bank name is required."
            )

            return redirect('bank_create')

        # =====================================================
        # PREVENT DUPLICATE BANK
        # =====================================================

        if Bank.objects.filter(
            company_id=selected_company_id,
            name__iexact=name,
            is_active=True
        ).exists():

            messages.error(
                request,
                "Bank with this name already exists."
            )

            return redirect('bank_create')

        # =====================================================
        # CREATE BANK
        # =====================================================

        bank = Bank.objects.create(

            company_id=selected_company_id,

            name=name,

            opening_balance=opening_balance,

            available_balance=opening_balance
        )

        # =====================================================
        # ACTIVITY LOG
        # =====================================================
        company = Company.objects.get(
            id=selected_company_id
        )

        log_activity(

        request=request,

        action="Created Bank",

        action_type='Created',

        description=f"""
    Bank Created Successfully

    Bank Name:
    {bank.name}

    Opening Balance:
    Rs. {bank.opening_balance}

    Available Balance:
    Rs. {bank.available_balance}

    Company name:
    {company.name}
    """
    )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Bank created successfully."
        )

        return redirect('bank_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'bank/create.html'
    )





from django.shortcuts import get_object_or_404


@login_required(login_url='login')
def bank_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH BANK FROM SELECTED COMPANY
    # =====================================================

    bank = get_object_or_404(
        Bank,
        pk=pk,
        company_id=selected_company_id
    )

    # =====================================================
    # UPDATE BANK
    # =====================================================

    if request.method == 'POST':

        # Store old values for activity log
        old_name = bank.name
        old_opening_balance = bank.opening_balance
        old_available_balance = bank.available_balance

        name = request.POST.get(
            'name',
            ''
        ).strip()

        opening_balance = Decimal(
            request.POST.get('opening_balance') or 0
        )

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Bank name is required."
            )

            return redirect(
                'bank_update',
                pk=bank.id
            )

        # =====================================================
        # PREVENT DUPLICATE BANK
        # =====================================================

        if Bank.objects.filter(
            company_id=selected_company_id,
            name__iexact=name,
            is_active=True
        ).exclude(id=bank.id).exists():

            messages.error(
                request,
                "Another bank with this name already exists."
            )

            return redirect(
                'bank_update',
                pk=bank.id
            )

        # =====================================================
        # UPDATE BANK
        # =====================================================

        bank.name = name
        bank.opening_balance = opening_balance

        bank.save(
            update_fields=[
                'name',
                'opening_balance'
            ]
        )

        # =====================================================
        # RECALCULATE BALANCE
        # =====================================================

        bank.recalculate_balance()

        # Refresh updated balance
        bank.refresh_from_db()

        # =====================================================
        # ACTIVITY LOG
        # =====================================================
        company = Company.objects.get(
            id=selected_company_id
        )
                
        log_activity(

        request=request,

        action="Updated Bank",

        action_type='Updated',

        description=f"""
    Bank Updated Successfully

    Old Bank Name:
    {old_name}

    New Bank Name:
    {bank.name}

    Old Opening Balance:
    Rs. {old_opening_balance}

    New Opening Balance:
    Rs. {bank.opening_balance}

    Old Available Balance:
    Rs. {old_available_balance}

    New Available Balance:
    Rs. {bank.available_balance}

    Company:
    {company.name}
    """
    )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Bank updated successfully."
        )

        return redirect('bank_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'bank/update.html',
        {
            'bank': bank
        }
    )




@login_required(login_url='login')
def bank_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH BANK FROM SELECTED COMPANY
    # =====================================================

    bank = get_object_or_404(
        Bank,
        pk=pk,
        company_id=selected_company_id
    )

    # =====================================================
    # DELETE BANK
    # =====================================================

    if request.method == 'POST':

        # Store values before delete
        bank_name = bank.name
        opening_balance = bank.opening_balance
        available_balance = bank.available_balance

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        company = Company.objects.get(
            id=selected_company_id
        )

        log_activity(

        request=request,

        action="Deleted Bank",

        action_type='Deleted',

        description=f"""
    Bank Deleted Successfully

    Bank Name:
    {bank_name}

    Opening Balance:
    Rs. {opening_balance}

    Available Balance:
    Rs. {available_balance}

    Company:
    {company.name}
    """
    )
        

        # =====================================================
        # DELETE BANK
        # =====================================================

        bank.delete()

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Bank deleted successfully."
        )

        return redirect('bank_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'bank/delete.html',
        {
            'bank': bank
        }
    )


from django.shortcuts import redirect, get_object_or_404

def toggle_bank_status(request, pk):
    bank = get_object_or_404(
        Bank,
        pk=pk,
        company_id=request.session.get('selected_company_id')
    )

    bank.is_active = not bank.is_active
    bank.save()

    return redirect('bank_index')


from itertools import chain
from django.db.models import F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404, render, redirect
from django.utils.dateparse import parse_date


def bank_log(request, pk):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    bank = get_object_or_404(Bank, pk=pk)

    client_id = request.GET.get('client')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # =========================
    # 📊 BASE QUERYSETS (COMPANY AWARE)
    # =========================
    payments = Payment.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client')

    expenses = Expense.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client', 'category')

    # =========================
    # 👤 CLIENT FILTER
    # =========================
    if client_id:
        payments = payments.filter(client_id=client_id)
        expenses = expenses.filter(client_id=client_id)

    # =========================
    # 📅 DATE FILTERS
    # =========================
    if start_date:
        sd = parse_date(start_date)
        payments = payments.filter(payment_date__gte=sd)
        expenses = expenses.filter(expense_date__gte=sd)

    if end_date:
        ed = parse_date(end_date)
        payments = payments.filter(payment_date__lte=ed)
        expenses = expenses.filter(expense_date__lte=ed)

    # =========================
    # 🔵 PAYMENTS → CREDIT
    # =========================
    payment_rows = payments.annotate(
        txn_date=F('payment_date'),
        txn_type=Value('Payment'),
        txn_description=Value('Client Payment'),
        category_name=Value('—'),
        credit=F('amount'),
        debit=Value(
            None,
            output_field=DecimalField(max_digits=12, decimal_places=2)
        ),
        client_name=F('client__name'),
    ).values(
        'txn_date',
        'txn_type',
        'txn_description',
        'category_name',
        'credit',
        'debit',
        'client_name',
    )

    # =========================
    # 🔴 EXPENSES → DEBIT
    # =========================
    expense_rows = expenses.annotate(
        txn_date=F('expense_date'),
        txn_type=Value('Spend'),
        txn_description=F('description'),
        category_name=Coalesce(F('category__name'), Value('—')),
        credit=Value(
            None,
            output_field=DecimalField(max_digits=12, decimal_places=2)
        ),
        debit=F('amount'),
        client_name=F('client__name'),
    ).values(
        'txn_date',
        'txn_type',
        'txn_description',
        'category_name',
        'credit',
        'debit',
        'client_name',
    )

    # =========================
    # 🔗 MERGE + SORT
    # =========================
    transactions = sorted(
        chain(payment_rows, expense_rows),
        key=lambda x: x['txn_date']
    )

    # =========================
    # 💰 RUNNING BALANCE
    # =========================
    balance = bank.opening_balance

    for row in transactions:
        if row['credit']:
            balance += row['credit']
        if row['debit']:
            balance -= row['debit']
        row['balance'] = balance

    return render(request, 'bank/bank_log.html', {
        'bank': bank,
        'logs': transactions,

        # 👇 only company clients
        'clients': Client.objects.filter(
            company_id=selected_company_id
        ),

        'selected_client': client_id,
        'start_date': start_date,
        'end_date': end_date,
    })





from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from decimal import Decimal
from django.utils.dateparse import parse_date


def clean_param(value):
    return value if value not in [None, '', 'None'] else None


def bank_log_pdf(request, pk):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    bank = get_object_or_404(Bank, pk=pk)

    # ✅ CLEAN QUERY PARAMS
    client_id = clean_param(request.GET.get('client'))
    start_date = clean_param(request.GET.get('start_date'))
    end_date = clean_param(request.GET.get('end_date'))

    # =========================
    # 📊 BASE QUERYSETS (COMPANY AWARE)
    # =========================
    payments = Payment.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client')

    expenses = Expense.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client', 'category')

    # 👤 CLIENT FILTER
    if client_id:
        payments = payments.filter(client_id=int(client_id))
        expenses = expenses.filter(client_id=int(client_id))

    # 📅 DATE FILTERS
    if start_date:
        sd = parse_date(start_date)
        payments = payments.filter(payment_date__gte=sd)
        expenses = expenses.filter(expense_date__gte=sd)

    if end_date:
        ed = parse_date(end_date)
        payments = payments.filter(payment_date__lte=ed)
        expenses = expenses.filter(expense_date__lte=ed)

    # =========================
    # 🔁 NORMALIZE TRANSACTIONS
    # =========================
    rows = []

    for p in payments.order_by('payment_date', 'id'):
        rows.append({
            'date': p.payment_date.strftime('%d-%m-%Y'),
            'client': p.client.name,
            'type': 'Payment',
            'desc': 'Client Payment',
            'credit': p.amount,
            'debit': Decimal('0.00'),
        })

    for e in expenses.order_by('expense_date', 'id'):
        rows.append({
            'date': e.expense_date.strftime('%d-%m-%Y'),
            'client': e.client.name,
            'type': 'Spend',
            'desc': e.description,
            'credit': Decimal('0.00'),
            'debit': e.amount,
        })

    rows = sorted(rows, key=lambda x: x['date'])

    # =========================
    # 💰 RUNNING BALANCE
    # =========================
    balance = bank.opening_balance
    for r in rows:
        balance += r['credit']
        balance -= r['debit']
        r['balance'] = balance

    # =========================
    # 📄 PDF GENERATION (LANDSCAPE)
    # =========================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="{bank.name}_bank_log.pdf"'
    )

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),   # ✅ LANDSCAPE
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()
    elements = []

    # =========================
    # 🧾 HEADER
    # =========================
    elements.append(
        Paragraph(f"<b>Bank Statement – {bank.name}</b>", styles['Title'])
    )

    elements.append(
        Paragraph(
            f"Opening Balance: <b>Rs. {bank.opening_balance:.2f}</b>",
            styles['Normal']
        )
    )

    if start_date or end_date:
        elements.append(
            Paragraph(
                f"Period: {start_date or '—'} to {end_date or '—'}",
                styles['Italic']
            )
        )

    elements.append(Spacer(1, 14))

    # =========================
    # 📊 TABLE
    # =========================
    table_data = [[
        'Date',
        'Client',
        'Type',
        'Description',
        'Credit (Rs.)',
        'Debit (Rs.)',
        'Balance (Rs.)'
    ]]

    for r in rows:
        table_data.append([
            r['date'],
            r['client'],
            r['type'],
            r['desc'],
            f"{r['credit']:.2f}" if r['credit'] else '—',
            f"{r['debit']:.2f}" if r['debit'] else '—',
            f"{r['balance']:.2f}",
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[85, 120, 80, 220, 90, 90, 95]
    )

    table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F0FE')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
    ]))

    elements.append(table)

    doc.build(elements)
    return response




#export bank log as excel will be added here

def bank_log_excel(request, pk):

    # =========================
    # COMPANY SESSION
    # =========================

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    bank = get_object_or_404(Bank, pk=pk)

    client_id = clean_param(request.GET.get('client'))
    start_date = clean_param(request.GET.get('start_date'))
    end_date = clean_param(request.GET.get('end_date'))

    # =========================
    # QUERYSETS
    # =========================

    payments = Payment.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client')

    expenses = Expense.objects.filter(
        bank=bank,
        client__company_id=selected_company_id
    ).select_related('client', 'category')

    if client_id:
        payments = payments.filter(client_id=int(client_id))
        expenses = expenses.filter(client_id=int(client_id))

    if start_date:
        sd = parse_date(start_date)
        payments = payments.filter(payment_date__gte=sd)
        expenses = expenses.filter(expense_date__gte=sd)

    if end_date:
        ed = parse_date(end_date)
        payments = payments.filter(payment_date__lte=ed)
        expenses = expenses.filter(expense_date__lte=ed)

    # =========================
    # NORMALIZE DATA
    # =========================

    rows = []

    for p in payments.order_by('payment_date', 'id'):
        rows.append({
            'date': p.payment_date,
            'client': p.client.name,
            'type': 'Payment',
            'desc': 'Client Payment',
            'credit': p.amount,
            'debit': Decimal('0.00'),
        })

    for e in expenses.order_by('expense_date', 'id'):
        rows.append({
            'date': e.expense_date,
            'client': e.client.name,
            'type': 'Spend',
            'desc': e.description or '—',
            'credit': Decimal('0.00'),
            'debit': e.amount,
        })

    rows = sorted(rows, key=lambda x: x['date'])

    # =========================
    # RUNNING BALANCE
    # =========================

    balance = bank.opening_balance

    for r in rows:
        balance += r['credit']
        balance -= r['debit']
        r['balance'] = balance

    # =========================
    # EXCEL
    # =========================

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # =========================
    # STYLES
    # =========================

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    total_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(bold=True)

    red_font = Font(
        color="FF0000",
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    thin = Side(style='thin', color='CCCCCC')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =========================
    # TITLE
    # =========================

    ws.merge_cells('A1:G1')

    title = ws['A1']

    title.value = f"BANK STATEMENT - {bank.name}"

    title.font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    title.fill = header_fill
    title.alignment = center

    # =========================
    # INFO
    # =========================

    ws['A3'] = "Opening Balance"
    ws['B3'] = float(bank.opening_balance)

    ws['A4'] = "Period"

    if start_date or end_date:
        ws['B4'] = f"{start_date or '—'} to {end_date or '—'}"
    else:
        ws['B4'] = "All"

    ws['A3'].font = bold_font
    ws['A4'].font = bold_font

    # =========================
    # HEADERS
    # =========================

    headers = [
        'Date',
        'Client',
        'Type',
        'Description',
        'Credit',
        'Debit',
        'Balance'
    ]

    row = 6

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row, col_num)

        cell.value = header
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row += 1

    total_credit = Decimal('0.00')
    total_debit = Decimal('0.00')

    # =========================
    # DATA ROWS
    # =========================

    for r in rows:

        total_credit += r['credit']
        total_debit += r['debit']

        data = [
            r['date'].strftime('%d-%m-%Y'),
            r['client'],
            r['type'],
            r['desc'],
            float(r['credit']) if r['credit'] else '—',
            float(r['debit']) if r['debit'] else '—',
            float(r['balance']),
        ]

        for col_num, value in enumerate(data, 1):

            cell = ws.cell(row, col_num)

            cell.value = value
            cell.border = border

            if col_num >= 5:
                cell.alignment = right

        # Negative balance
        if r['balance'] < 0:
            ws.cell(row, 7).font = red_font

        row += 1

    # =========================
    # TOTAL ROW
    # =========================

    final_balance = balance

    totals = [
        'GRAND TOTAL',
        '',
        '',
        '',
        float(total_credit),
        float(total_debit),
        float(final_balance)
    ]

    for col_num, value in enumerate(totals, 1):

        cell = ws.cell(row, col_num)

        cell.value = value

        cell.font = white_font
        cell.fill = header_fill
        cell.border = border

        if col_num >= 5:
            cell.alignment = right

    if final_balance < 0:
        ws.cell(row, 7).font = red_font

    # =========================
    # COLUMN WIDTHS
    # =========================

    widths = {
        1: 15,
        2: 28,
        3: 16,
        4: 40,
        5: 18,
        6: 18,
        7: 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        f'attachment; filename="{bank.name}_bank_statement.xlsx"'
    )

    wb.save(response)

    return response







# bank to bank transfer views will be added here

# views.py
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404

def transfer_list(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    banks = Bank.objects.filter(company_id=selected_company_id)

    # 🔎 Filters
    search = request.GET.get('q', '')
    from_bank = request.GET.get('from_bank')
    to_bank = request.GET.get('to_bank')

    transfers = BankTransfer.objects.filter(
        from_bank__company_id=selected_company_id
    ).select_related('from_bank', 'to_bank').order_by('-id')

    if search:
        transfers = transfers.filter(
            Q(from_bank__name__icontains=search) |
            Q(to_bank__name__icontains=search)
        )

    if from_bank:
        transfers = transfers.filter(from_bank_id=from_bank)

    if to_bank:
        transfers = transfers.filter(to_bank_id=to_bank)

    return render(request, 'transfer/list.html', {
        'transfers': transfers,
        'banks': banks
    })




from decimal import Decimal
from django.db import transaction
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from .models import (
    Bank,
    BankTransfer,
    ActivityLog
)






@login_required(login_url='login')
def transfer_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH ACTIVE BANKS
    # =====================================================

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    )

    # =====================================================
    # CREATE TRANSFER
    # =====================================================

    if request.method == 'POST':

        from_bank_id = request.POST.get('from_bank')

        to_bank_id = request.POST.get('to_bank')

        # ✅ NEW DESCRIPTION FIELD
        description = request.POST.get(
            'description',
            ''
        ).strip()

        transfer_date = request.POST.get(
            'transfer_date'
        )

        amount_raw = request.POST.get(
            'amount'
        )

        # =====================================================
        # VALIDATION
        # =====================================================

        if not amount_raw:

            messages.error(
                request,
                "Amount is required"
            )

            return redirect('transfer_create')

        try:

            amount = Decimal(
                str(amount_raw)
                .replace(',', '')
                .strip()
            )

        except:

            messages.error(
                request,
                "Invalid amount format"
            )

            return redirect('transfer_create')

        if amount <= 0:

            messages.error(
                request,
                "Amount must be greater than zero"
            )

            return redirect('transfer_create')

        # =====================================================
        # FETCH BANKS
        # =====================================================

        from_bank = banks.filter(
            id=from_bank_id
        ).first()

        to_bank = banks.filter(
            id=to_bank_id
        ).first()

        if not from_bank or not to_bank:

            messages.error(
                request,
                "Invalid bank selection"
            )

            return redirect('transfer_create')

        # =====================================================
        # PREVENT SAME BANK TRANSFER
        # =====================================================

        if from_bank.id == to_bank.id:

            messages.error(
                request,
                "Cannot transfer to the same bank"
            )

            return redirect('transfer_create')

        # =====================================================
        # PREVENT OVERDRAFT
        # =====================================================

        if from_bank.available_balance < amount:

            messages.error(
                request,
                f"Insufficient balance in {from_bank.name}"
            )

            return redirect('transfer_create')

        # =====================================================
        # SAVE TRANSFER
        # =====================================================

        with transaction.atomic():

            transfer = BankTransfer.objects.create(

                from_bank=from_bank,

                to_bank=to_bank,

                amount=amount,

                # ✅ SAVE DESCRIPTION
                description=description,

                transfer_date=transfer_date
            )

            # =====================================================
            # RECALCULATE BALANCES
            # =====================================================

            recalculate_banks(
                from_bank,
                to_bank
            )

            # Refresh latest balances
            from_bank.refresh_from_db()

            to_bank.refresh_from_db()


            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

    request=request,

    action="Created Bank Transfer",

    action_type='Created',

    description=f"""
Bank Transfer Created Successfully

Transfer ID:
{transfer.id}

From Bank:
{from_bank.name}

To Bank:
{to_bank.name}

Transfer Amount:
Rs. {amount}

Transfer Note:
{description or 'N/A'}

Transfer Date:
{transfer_date}

From Bank Balance After Transfer:
Rs. {from_bank.available_balance}

To Bank Balance After Transfer:
Rs. {to_bank.available_balance}

Company:
{from_bank.company.name}
"""
)
            


        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Transfer created successfully"
        )

        return redirect('transfer_list')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'transfer/create.html',
        {
            'today_date': timezone.now().date().strftime('%Y-%m-%d'),
            'banks': banks
        }
    )





@login_required(login_url='login')
def transfer_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH TRANSFER
    # =====================================================

    transfer = get_object_or_404(

        BankTransfer,

        pk=pk,

        from_bank__company_id=selected_company_id
    )

    # =====================================================
    # DELETE TRANSFER
    # =====================================================

    if request.method == 'POST':

        with transaction.atomic():

            from_bank = transfer.from_bank

            to_bank = transfer.to_bank

            amount = transfer.amount

            transfer_date = transfer.transfer_date

            description = transfer.description

            transfer_id = transfer.id

            # =====================================================
            # STORE OLD BALANCES
            # =====================================================

            old_from_balance = from_bank.available_balance

            old_to_balance = to_bank.available_balance

            # =====================================================
            # DELETE TRANSFER
            # =====================================================

            transfer.delete()

            # =====================================================
            # RECALCULATE BANKS
            # =====================================================

            recalculate_banks(
                from_bank,
                to_bank
            )

            # Refresh latest balances
            from_bank.refresh_from_db()

            to_bank.refresh_from_db()

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

        request=request,

        action="Deleted Bank Transfer",

        action_type='Deleted',

        description=f"""
    Bank Transfer Deleted Successfully

    Transfer ID:
    {transfer_id}

    From Bank:
    {from_bank.name}

    To Bank:
    {to_bank.name}

    Transfer Amount:
    Rs. {amount}

    Transfer Note:
    {description or 'N/A'}

    Transfer Date:
    {transfer_date}

    Old From Bank Balance:
    Rs. {old_from_balance}

    New From Bank Balance:
    Rs. {from_bank.available_balance}

    Old To Bank Balance:
    Rs. {old_to_balance}

    New To Bank Balance:
    Rs. {to_bank.available_balance}

    Company:
    {from_bank.company.name}
    """
    )
        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Transfer deleted successfully"
        )

        return redirect('transfer_list')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'transfer/delete.html',
        {
            'transfer': transfer
        }
    )



#cash views will be added here
from django.utils import timezone
from zoneinfo import ZoneInfo


from django.utils import timezone
from zoneinfo import ZoneInfo
from django.utils.dateparse import parse_date
from django.db.models import Sum

def cash_index(request):
    cash_list = Cash.objects.select_related('client')

    # 🔎 FILTER INPUTS
    filter_type = request.GET.get('filter_type')
    filter_date = request.GET.get('date')
    filter_month = request.GET.get('month')
    filter_year = request.GET.get('year')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    client_id = request.GET.get('client')
    cash_type = request.GET.get('cash_type')  # in / out

    # 👤 CLIENT FILTER
    if client_id:
        cash_list = cash_list.filter(client_id=client_id)

    # 🔄 CASH TYPE FILTER
    if cash_type in [Cash.CASH_IN, Cash.CASH_OUT]:
        cash_list = cash_list.filter(cash_type=cash_type)

    # 📅 DATE RANGE FILTER (PRIORITY)
    if start_date:
        cash_list = cash_list.filter(
            cash_date__gte=parse_date(start_date)
        )

    if end_date:
        cash_list = cash_list.filter(
            cash_date__lte=parse_date(end_date)
        )

    # 📆 DAY / MONTH / YEAR FILTERS (OPTIONAL)
    if filter_type == 'day' and filter_date:
        cash_list = cash_list.filter(cash_date=parse_date(filter_date))

    elif filter_type == 'month' and filter_month and filter_year:
        cash_list = cash_list.filter(
            cash_date__month=int(filter_month),
            cash_date__year=int(filter_year)
        )

    elif filter_type == 'year' and filter_year:
        cash_list = cash_list.filter(
            cash_date__year=int(filter_year)
        )

    cash_list = cash_list.order_by('-cash_date')

    # 💰 TOTAL CASH (FILTERED RESULT)
    cash_in = cash_list.filter(
        cash_type=Cash.CASH_IN
    ).aggregate(total=Sum('amount'))['total'] or 0

    cash_out = cash_list.filter(
        cash_type=Cash.CASH_OUT
    ).aggregate(total=Sum('amount'))['total'] or 0

    total_cash = cash_in - cash_out

    return render(request, 'cash/index.html', {
        'cash_list': cash_list,
        'total_cash': total_cash,

        # filters
        'filter_type': filter_type,
        'filter_date': filter_date,
        'filter_month': filter_month,
        'filter_year': filter_year,
        'start_date': start_date,
        'end_date': end_date,
        'selected_client': client_id,
        'selected_cash_type': cash_type,

        # dropdown data
        'clients': Client.objects.all(),
        'months': range(1, 13),
        'current_year': timezone.now().year,
    })



def cash_create(request):
    clients = Client.objects.all()

    if request.method == 'POST':
        client_id = request.POST.get('client')
        amount = Decimal(request.POST.get('amount'))
        cash_type = request.POST.get('cash_type')
        description = request.POST.get('description')
        cash_date = request.POST.get('cash_date')

        if cash_type not in [Cash.CASH_IN, Cash.CASH_OUT]:
            messages.error(request, "Invalid cash type")
            return render(request, 'cash/create.html', {
                'clients': clients
            })

        with transaction.atomic():

            # 🔴 VALIDATE CASH OUT
            if cash_type == Cash.CASH_OUT:
                cash_in = Cash.objects.filter(
                    cash_type=Cash.CASH_IN
                ).aggregate(total=Sum('amount'))['total'] or 0

                cash_out = Cash.objects.filter(
                    cash_type=Cash.CASH_OUT
                ).aggregate(total=Sum('amount'))['total'] or 0

                if cash_in - cash_out < amount:
                    messages.error(request, "Insufficient cash balance")
                    return render(request, 'cash/create.html', {
                        'clients': clients
                    })

            Cash.objects.create(
                client_id=client_id,
                amount=amount,
                cash_type=cash_type,
                description=description,
                cash_date=cash_date
            )

        return redirect('cash_index')

    return render(request, 'cash/create.html', {
        'clients': clients
    })



def cash_update(request, pk):
    cash = get_object_or_404(Cash, pk=pk)
    clients = Client.objects.all()

    old_amount = cash.amount
    old_type = cash.cash_type

    if request.method == 'POST':
        client_id = request.POST.get('client')
        new_amount = Decimal(request.POST.get('amount'))
        new_type = request.POST.get('cash_type')
        description = request.POST.get('description')
        cash_date = request.POST.get('cash_date')

        with transaction.atomic():

            # 🔁 REVERSE OLD EFFECT
            if old_type == Cash.CASH_IN:
                reverse_type = Cash.CASH_OUT
            else:
                reverse_type = Cash.CASH_IN

            Cash.objects.create(
                client=cash.client,
                amount=old_amount,
                cash_type=reverse_type,
                description="Cash update reversal",
                cash_date=cash.cash_date
            )

            # 🔴 VALIDATE NEW CASH OUT
            if new_type == Cash.CASH_OUT:
                cash_in = Cash.objects.filter(
                    cash_type=Cash.CASH_IN
                ).aggregate(total=Sum('amount'))['total'] or 0

                cash_out = Cash.objects.filter(
                    cash_type=Cash.CASH_OUT
                ).aggregate(total=Sum('amount'))['total'] or 0

                if cash_in - cash_out < new_amount:
                    messages.error(request, "Insufficient cash balance")
                    return render(request, 'cash/update.html', {
                        'cash': cash,
                        'clients': clients
                    })

            # 🔁 APPLY NEW ENTRY
            Cash.objects.create(
                client_id=client_id,
                amount=new_amount,
                cash_type=new_type,
                description=description,
                cash_date=cash_date
            )

            # ❌ DELETE OLD ROW (AUDIT PRESERVED VIA REVERSAL)
            cash.delete()

        return redirect('cash_index')

    return render(request, 'cash/update.html', {
        'cash': cash,
        'clients': clients
    })



def cash_delete(request, pk):
    cash = get_object_or_404(Cash, pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            cash.delete()
        return redirect('cash_index')

    return render(request, 'cash/delete.html', {
        'cash': cash
    })




# available balance views will be added here




from django.db.models import Sum, Max, Q
from django.shortcuts import render, redirect
from decimal import Decimal
from .models import Bank


def available_amount(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    # =========================
    # 💵 CASH BANK
    # =========================
    cash_bank = Bank.objects.filter(
        company_id=selected_company_id,
        name__iexact='cash'
    ).first()

    total_cash = Decimal('0.00')
    last_cash_date = None

    if cash_bank:

        # ✅ ALWAYS use model logic
        total_cash = cash_bank.calculated_balance()

        last_cash_date = max(
            filter(None, [
                cash_bank.payments.aggregate(d=Max('payment_date'))['d'],
                cash_bank.expenses.aggregate(d=Max('expense_date'))['d'],
                cash_bank.transfers_in.aggregate(d=Max('transfer_date'))['d'],
                cash_bank.transfers_out.aggregate(d=Max('transfer_date'))['d'],
            ]),
            default=None
        )

    # =========================
    # 🏦 ALL OTHER BANKS (FIXED)
    # =========================
    cheque_banks = Bank.objects.filter(
        company_id=selected_company_id,is_active=True
    ).exclude(name__iexact='cash')

    total_bank = Decimal('0.00')

    for bank in cheque_banks:

        # ✅ use model calculation (includes transfers also)
        bank.available_balance = bank.calculated_balance()

        # 🗓 LAST ACTIVITY
        bank.last_transaction_date = max(
            filter(None, [
                bank.payments.aggregate(d=Max('payment_date'))['d'],
                bank.expenses.aggregate(d=Max('expense_date'))['d'],
                bank.transfers_in.aggregate(d=Max('transfer_date'))['d'],
                bank.transfers_out.aggregate(d=Max('transfer_date'))['d'],
            ]),
            default=None
        )

        total_bank += bank.available_balance

    return render(request, 'accounts/available_amount.html', {
        'total_cash': total_cash,
        'last_cash_date': last_cash_date,
        'cheque_banks': cheque_banks,
        'total_bank': total_bank,
    })





#payment 



from django.http import JsonResponse
from decimal import Decimal
from .models import Payment


def check_payment_duplicate(request):

    selected_company_id = request.session.get('selected_company_id')

    client_id = request.GET.get('client')
    bank_id = request.GET.get('bank')
    payment_date = request.GET.get('payment_date')
    amount = request.GET.get('amount')

    # Required fields
    if not (client_id and payment_date and amount):
        return JsonResponse({"duplicate": False})

    try:
        amount = Decimal(str(amount).replace(',', '').strip())
    except:
        return JsonResponse({"duplicate": False})

    duplicate = Payment.objects.filter(
        client_id=client_id,
        client__company_id=selected_company_id,
        bank_id=bank_id,
        payment_date=payment_date,
        amount=amount
    ).first()

    if duplicate:
        return JsonResponse({
            "duplicate": True,
            "client": duplicate.client.name,
            "amount": str(duplicate.amount),
            "date": str(duplicate.payment_date),
            "bank": duplicate.bank.name
        })

    return JsonResponse({"duplicate": False})



from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date

from .models import Payment, Client, Bank


def payment_index(request):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_id = request.GET.get('client')
    mode = request.GET.get('mode')        # cash / cheque
    bank_id = request.GET.get('bank')

    # =========================
    # 📊 BASE QUERYSET (COMPANY AWARE)
    # =========================
    payments = Payment.objects.filter(
        client__company_id=selected_company_id
    ).select_related('client', 'bank')

    # =========================
    # 📅 DATE FILTERS
    # =========================
    if start_date:
        payments = payments.filter(
            payment_date__gte=parse_date(start_date)
        )

    if end_date:
        payments = payments.filter(
            payment_date__lte=parse_date(end_date)
        )

    # =========================
    # 👤 CLIENT FILTER
    # =========================
    if client_id:
        payments = payments.filter(client_id=client_id)

    # =========================
    # 💳 MODE FILTER
    # =========================
    if mode in ['cash', 'cheque']:
        payments = payments.filter(payment_mode=mode)

    # =========================
    # 🏦 BANK FILTER
    # =========================
    if bank_id:
        payments = payments.filter(bank_id=bank_id)

    payments = payments.order_by('-payment_date', '-id')

    # =========================
    # 📥 DROPDOWN DATA (COMPANY AWARE)
    # =========================
    clients = Client.objects.filter(
        company_id=selected_company_id
    )

    cheque_banks = Bank.objects.filter(
        payments__client__company_id=selected_company_id
    ).exclude(
        name__iexact='cash'
    ).distinct()

    cash_bank = Bank.objects.filter(name__iexact='cash').first()

    return render(request, 'payment/index.html', {
        'payments': payments,

        'clients': clients,
        'banks': cheque_banks,
        'cash_bank': cash_bank,

        'start_date': start_date,
        'end_date': end_date,
        'selected_client': client_id,
        'selected_mode': mode,
        'selected_bank': bank_id,
    })









# def payment_create(request):

#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         return redirect('dashboard')

#     clients = Client.objects.filter(company_id=selected_company_id,is_active=True)
#     banks = Bank.objects.filter(
#         company_id=selected_company_id,
#         is_active=True
#     ).order_by('name')


#     if request.method == 'POST':

#         client_id = request.POST.get('client')
#         bank_id = request.POST.get('bank')
#         amount = Decimal(request.POST.get('amount'))
#         payment_mode = request.POST.get('payment_mode')
#         payment_date = request.POST.get('payment_date')

#         client = Client.objects.get(
#             id=client_id,
#             company_id=selected_company_id,is_active=True
#         )

#         # 🔥 Get total already paid
#         total_paid = client.payments.aggregate(
#             total=Sum('amount')
#         )['total'] or Decimal('0.00')

#         new_total = total_paid + amount

#         # ❌ VALIDATION
#         if new_total > client.budget:
#             messages.error(
#                 request,
#                 f"❌ Payment exceeds project value! "
#                 f"Remaining allowed: ₹ {client.budget - total_paid:,.2f}"
#             )
#             return render(request, 'payment/create.html', {
#                 'clients': clients,
#                 'banks': banks
#             })

#         with transaction.atomic():

#             # 💵 CASH MODE
#             if payment_mode == 'cash':
#                 cash_bank = Bank.objects.get(
#                     name__iexact='cash',
#                     company_id=selected_company_id
#                 )

#                 Payment.objects.create(
#                     client=client,
#                     bank=cash_bank,
#                     amount=amount,
#                     payment_mode=Payment.CASH,
#                     payment_date=payment_date
#                 )

#                 cash_bank.recalculate_balance()

#             # 🔵 CHEQUE MODE
#             elif payment_mode == 'cheque':

#                 if not bank_id:
#                     messages.error(request, "Select a bank for cheque payment")
#                     return render(request, 'payment/create.html', {
#                         'clients': clients,
#                         'banks': banks
#                     })

#                 bank = Bank.objects.get(
#                     id=bank_id,
#                     company_id=selected_company_id,is_active=True
#                 )

#                 Payment.objects.create(
#                     client=client,
#                     bank=bank,
#                     amount=amount,
#                     payment_mode=Payment.CHEQUE,
#                     payment_date=payment_date
#                 )

#                 bank.recalculate_balance()

#         messages.success(request, "✅ Payment added successfully")
#         return redirect('payment_index')

#     return render(request, 'payment/create.html', {
#         'clients': clients,
#         'banks': banks
#     })



from decimal import Decimal
from django.db import transaction
from django.db.models import Sum
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .utils import *
from .models import (
    Client,
    Bank,
    Payment,
    ActivityLog
)


@login_required(login_url='login')
def payment_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH CLIENTS & BANKS
    # =====================================================

    clients = Client.objects.filter(
        company_id=selected_company_id,
        is_active=True
    )

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    ).order_by('name')

    # =====================================================
    # CREATE PAYMENT
    # =====================================================

    if request.method == 'POST':

        client_id = request.POST.get('client')

        bank_id = request.POST.get('bank')

        payment_mode = request.POST.get('payment_mode')

        payment_date = request.POST.get('payment_date')

        # =====================================================
        # VALIDATE AMOUNT
        # =====================================================

        try:

            amount = Decimal(
                request.POST.get('amount')
            )

        except:

            messages.error(
                request,
                "Invalid payment amount"
            )

            return render(
                request,
                'payment/create.html',
                {
                    'clients': clients,
                    'banks': banks
                }
            )

        # =====================================================
        # FETCH CLIENT
        # =====================================================

        client = Client.objects.get(
            id=client_id,
            company_id=selected_company_id,
            is_active=True
        )

        # =====================================================
        # PROJECT BUDGET VALIDATION
        # =====================================================

        total_paid = client.payments.aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        new_total = total_paid + amount

        if new_total > client.budget:

            messages.error(
                request,
                f"❌ Payment exceeds project value! "
                f"Remaining allowed: ₹ {client.budget - total_paid:,.2f}"
            )

            return render(
                request,
                'payment/create.html',
                {
                    'clients': clients,
                    'banks': banks
                }
            )

        # =====================================================
        # SAVE PAYMENT
        # =====================================================

        with transaction.atomic():

            # =====================================================
            # CASH PAYMENT
            # =====================================================

            if payment_mode == 'cash':

                cash_bank = Bank.objects.get(
                    name__iexact='cash',
                    company_id=selected_company_id
                )

                payment = Payment.objects.create(

                    client=client,

                    bank=cash_bank,

                    amount=amount,

                    payment_mode=Payment.CASH,

                    payment_date=payment_date
                )

                cash_bank.recalculate_balance()

                cash_bank.refresh_from_db()

                # =====================================================
                # ACTIVITY LOG
                # =====================================================

                log_activity(

                request=request,

                action="Created Cash Payment",

                action_type='Created',

                description=f"""
            Cash Payment Added Successfully

            Payment ID:
            {payment.id}

            Client:
            {client.name}

            Location:
            {client.location or 'N/A'}

            Amount:
            Rs. {amount}

            Payment Date:
            {payment_date}

            Payment Mode:
            Cash

            Bank:
            {cash_bank.name}

            Updated Bank Balance:
            Rs. {cash_bank.available_balance}

            Company:
            {client.company.name}
            """
            )

            # =====================================================
            # CHEQUE PAYMENT
            # =====================================================

            elif payment_mode == 'cheque':

                if not bank_id:

                    messages.error(
                        request,
                        "Select a bank for cheque payment"
                    )

                    return render(
                        request,
                        'payment/create.html',
                        {
                            'clients': clients,
                            'banks': banks
                        }
                    )

                bank = Bank.objects.get(
                    id=bank_id,
                    company_id=selected_company_id,
                    is_active=True
                )

                payment = Payment.objects.create(

                    client=client,

                    bank=bank,

                    amount=amount,

                    payment_mode=Payment.CHEQUE,

                    payment_date=payment_date
                )

                bank.recalculate_balance()

                bank.refresh_from_db()

                # =====================================================
                # ACTIVITY LOG
                # =====================================================

                log_activity(

            request=request,

            action="Created Cheque Payment",

            action_type='Created',

            description=f"""
        Cheque Payment Added Successfully

        Payment ID:
        {payment.id}

        Client:
        {client.name}

        Location:
        {client.location or 'N/A'}

        Amount:
        Rs. {amount}

        Payment Date:
        {payment_date}

        Payment Mode:
        Cheque

        Bank:
        {bank.name}

        Updated Bank Balance:
        Rs. {bank.available_balance}

        Company:
        {client.company.name}
        """
        )
        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "✅ Payment added successfully"
        )

        return redirect('payment_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'payment/create.html',
        {
            'clients': clients,
            'banks': banks
        }
    )









@login_required(login_url='login')
def payment_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH PAYMENT
    # =====================================================

    payment = get_object_or_404(
        Payment,
        pk=pk,
        client__company_id=selected_company_id
    )

    # =====================================================
    # FETCH CLIENTS & BANKS
    # =====================================================

    clients = Client.objects.filter(
        company_id=selected_company_id,
        is_active=True
    )

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    ).order_by('name')

    # Store old values for activity log
    old_client = payment.client
    old_bank = payment.bank
    old_amount = payment.amount
    old_mode = payment.payment_mode
    old_date = payment.payment_date

    # =====================================================
    # UPDATE PAYMENT
    # =====================================================

    if request.method == 'POST':

        client_id = request.POST.get('client')

        bank_id = request.POST.get('bank')

        new_mode = request.POST.get('payment_mode')

        new_date = request.POST.get('payment_date')

        # =====================================================
        # VALIDATE AMOUNT
        # =====================================================

        try:

            new_amount = Decimal(
                request.POST.get('amount')
            )

        except:

            messages.error(
                request,
                "Invalid payment amount"
            )

            return render(
                request,
                'payment/update.html',
                {
                    'payment': payment,
                    'clients': clients,
                    'banks': banks
                }
            )

        # =====================================================
        # FETCH CLIENT
        # =====================================================

        client = get_object_or_404(
            Client,
            id=client_id,
            company_id=selected_company_id
        )

        # =====================================================
        # BUDGET VALIDATION
        # =====================================================

        total_paid = client.payments.exclude(
            id=payment.id
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0.00')

        new_total = total_paid + new_amount

        if new_total > client.budget:

            remaining_allowed = (
                client.budget - total_paid
            )

            messages.error(
                request,
                f"❌ Payment exceeds project value! "
                f"Remaining allowed: ₹ {remaining_allowed:,.2f}"
            )

            return render(
                request,
                'payment/update.html',
                {
                    'payment': payment,
                    'clients': clients,
                    'banks': banks
                }
            )

        # =====================================================
        # UPDATE PAYMENT
        # =====================================================

        with transaction.atomic():

            # =====================================================
            # STORE OLD BANK
            # =====================================================

            previous_bank = payment.bank

            # =====================================================
            # CASH MODE
            # =====================================================

            if new_mode == Payment.CASH:

                cash_bank = Bank.objects.get(
                    name__iexact='cash',
                    company_id=selected_company_id
                )

                payment.bank = cash_bank

                payment.payment_mode = Payment.CASH

            # =====================================================
            # CHEQUE MODE
            # =====================================================

            elif new_mode == Payment.CHEQUE:

                if not bank_id:

                    messages.error(
                        request,
                        "Select bank for cheque payment"
                    )

                    return render(
                        request,
                        'payment/update.html',
                        {
                            'payment': payment,
                            'clients': clients,
                            'banks': banks
                        }
                    )

                bank = Bank.objects.get(
                    id=bank_id,
                    company_id=selected_company_id,
                    is_active=True
                )

                payment.bank = bank

                payment.payment_mode = Payment.CHEQUE

            # =====================================================
            # UPDATE PAYMENT FIELDS
            # =====================================================

            payment.client = client

            payment.amount = new_amount

            payment.payment_date = new_date

            payment.save()

            # =====================================================
            # RECALCULATE BANKS
            # =====================================================

            recalculate_banks(
                previous_bank,
                payment.bank
            )

            # Refresh latest balance
            payment.bank.refresh_from_db()

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

            request=request,

            action="Updated Payment",

            action_type='Updated',

            description=f"""
        Payment Updated Successfully

        Payment ID:
        {payment.id}

        Old Client:
        {old_client.name}

        New Client:
        {client.name}

        Old Amount:
        Rs. {old_amount}

        New Amount:
        Rs. {new_amount}

        Old Payment Mode:
        {old_mode}

        New Payment Mode:
        {new_mode}

        Old Bank:
        {old_bank.name if old_bank else 'N/A'}

        New Bank:
        {payment.bank.name}

        Old Payment Date:
        {old_date}

        New Payment Date:
        {new_date}

        Updated Bank Balance:
        Rs. {payment.bank.available_balance}

        Company:
        {client.company.name}
        """
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "✅ Payment updated successfully"
        )

        return redirect('payment_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'payment/update.html',
        {
            'payment': payment,
            'clients': clients,
            'banks': banks
        }
    )

@login_required(login_url='login')
def payment_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH PAYMENT
    # =====================================================

    payment = get_object_or_404(
        Payment,
        pk=pk,
        client__company_id=selected_company_id
    )

    bank = payment.bank

    # =====================================================
    # STORE VALUES BEFORE DELETE
    # =====================================================

    client_name = payment.client.name

    client_location = payment.client.location

    amount = payment.amount

    payment_mode = payment.payment_mode

    payment_date = payment.payment_date

    bank_name = bank.name if bank else 'N/A'

    payment_id = payment.id

    # =====================================================
    # DELETE PAYMENT
    # =====================================================

    if request.method == 'POST':

        with transaction.atomic():

            # =====================================================
            # STORE OLD BANK BALANCE
            # =====================================================

            old_balance = (

                bank.available_balance

                if bank

                else Decimal('0.00')
            )

            # =====================================================
            # DELETE PAYMENT
            # =====================================================

            payment.delete()

            # =====================================================
            # RECALCULATE BANK
            # =====================================================

            recalculate_banks(bank)

            # Refresh latest balance
            if bank:

                bank.refresh_from_db()

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

        request=request,

        action="Deleted Payment",

        action_type='Deleted',

        description=f"""
    Payment Deleted Successfully

    Payment ID:
    {payment_id}

    Client:
    {client_name}

    Location:
    {client_location or 'N/A'}

    Amount:
    Rs. {amount}

    Payment Mode:
    {payment_mode}

    Payment Date:
    {payment_date}

    Bank:
    {bank_name}

    Old Bank Balance:
    Rs. {old_balance}

    New Bank Balance:
    Rs. {bank.available_balance if bank else '0.00'}

    Company:
    {payment.client.company.name}
    """
    )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Payment deleted successfully."
        )

        return redirect('payment_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'payment/delete.html',
        {
            'payment': payment
        }
    )



#worker views will be added here



def worker_index(request):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    # 👷 Workers only for selected company
    workers = (
        Worker.objects
        .filter(company_id=selected_company_id)
        .order_by('name')
    )

    return render(request, 'worker/index.html', {
        'workers': workers
    })



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages



@login_required(login_url='login')
def worker_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    company_id = request.session.get(
        'selected_company_id'
    )

    if not company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=company_id
    )

    # =====================================================
    # CREATE WORKER
    # =====================================================

    if request.method == 'POST':

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Worker name is required"
            )

        else:

            # =====================================================
            # PREVENT DUPLICATE WORKER
            # =====================================================

            if Worker.objects.filter(
                company_id=company_id,
                name__iexact=name
            ).exists():

                messages.warning(
                    request,
                    "This worker Team already exists for this company."
                )

                return redirect('worker_create')

            # =====================================================
            # CREATE WORKER
            # =====================================================

            worker = Worker.objects.create(

                name=name,

                company_id=company_id
            )

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

            request=request,

            action="Created Worker",

            action_type='Created',

            description=f"""
        Worker Team Created Successfully

        Worker Team:
        {worker.name}

        Company:
        {company.name}

        Worker ID:
        {worker.id}
        """
        )

            # =====================================================
            # SUCCESS MESSAGE
            # =====================================================

            messages.success(
                request,
                "Worker added successfully"
            )

            return redirect('worker_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker/create.html',
        {
            'title': 'Add Worker',
            'company': company
        }
    )







@login_required(login_url='login')
def worker_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    company_id = request.session.get(
        'selected_company_id'
    )

    if not company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=company_id
    )

    # =====================================================
    # FETCH WORKER
    # =====================================================

    worker = get_object_or_404(
        Worker,
        pk=pk,
        company_id=company_id
    )

    # =====================================================
    # UPDATE WORKER
    # =====================================================

    if request.method == 'POST':

        # Store old value for activity log
        old_name = worker.name

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Worker name is required"
            )

        else:

            # =====================================================
            # PREVENT DUPLICATE WORKER
            # =====================================================

            duplicate = Worker.objects.filter(
                company_id=company_id,
                name__iexact=name
            ).exclude(
                pk=worker.pk
            ).exists()

            if duplicate:

                messages.warning(
                    request,
                    "Another worker Team with this name already exists."
                )

                return redirect(
                    'worker_update',
                    pk=pk
                )

            # =====================================================
            # UPDATE WORKER
            # =====================================================

            worker.name = name

            worker.company_id = company_id

            worker.save()

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

        request=request,

        action="Updated Worker",

        action_type='Updated',

        description=f"""
    Worker Team Updated Successfully

    Old Worker Team:
    {old_name}

    New Worker Team:
    {worker.name}

    Company:
    {company.name}

    Worker ID:
    {worker.id}
    """
    )

            # =====================================================
            # SUCCESS MESSAGE
            # =====================================================

            messages.success(
                request,
                "Worker updated successfully"
            )

            return redirect('worker_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker/update.html',
        {
            'title': 'Update Worker',
            'worker': worker,
            'company': company
        }
    )





@login_required(login_url='login')
def worker_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    company_id = request.session.get(
        'selected_company_id'
    )

    if not company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=company_id
    )

    # =====================================================
    # FETCH WORKER
    # =====================================================

    worker = get_object_or_404(
        Worker,
        pk=pk,
        company_id=company_id
    )

    # Store values before delete
    worker_name = worker.name
    worker_id = worker.id

    # =====================================================
    # DELETE WORKER
    # =====================================================

    if request.method == 'POST':

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        log_activity(

            request=request,

            action="Deleted Worker",

            action_type='Deleted',

            description=f"""
        Worker Team Deleted Successfully

        Worker Team:
        {worker_name}

        Worker ID:
        {worker_id}

        Company:
        {company.name}
        """
        )

        # =====================================================
        # DELETE WORKER
        # =====================================================

        worker.delete()

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Worker deleted successfully"
        )

        return redirect('worker_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker/delete.html',
        {
            'worker': worker
        }
    )


def worker_name_index(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    worker_id = request.GET.get('worker')  # string or None

    # 🏢 Workers of selected company
    workers = Worker.objects.filter(company_id=selected_company_id)

    # 👥 Worker names ONLY under selected company
    names = WorkerName.objects.select_related('worker').filter(
        worker__company_id=selected_company_id
    )

    # 🔎 Apply filter if worker selected
    if worker_id:
        names = names.filter(worker_id=worker_id)

    return render(request, 'worker_name/index.html', {
        'workers': workers,
        'worker_names': names,
        'selected_worker': worker_id,
    })





@login_required(login_url='login')
def worker_name_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH WORKERS
    # =====================================================

    workers = Worker.objects.filter(
        company_id=selected_company_id
    )

    # =====================================================
    # CREATE WORKER NAME
    # =====================================================

    if request.method == 'POST':

        worker_id = request.POST.get('worker')

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not worker_id or not name:

            messages.error(
                request,
                "Worker and name are required"
            )

            return redirect('worker_name_create')

        # =====================================================
        # FETCH WORKER
        # =====================================================

        worker = get_object_or_404(
            Worker,
            id=worker_id,
            company_id=selected_company_id
        )

        # =====================================================
        # PREVENT DUPLICATE WORKER NAME
        # =====================================================

        if WorkerName.objects.filter(
            worker=worker,
            name__iexact=name
        ).exists():

            messages.warning(
                request,
                "This worker name already exists for this worker Team."
            )

            return redirect('worker_name_create')

        # =====================================================
        # CREATE WORKER NAME
        # =====================================================

        worker_name = WorkerName.objects.create(

            worker=worker,

            name=name
        )

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Created Worker Name",

            description=f"""
Worker Name Created Successfully

Worker Team:
{worker.name}

Worker Name:
{worker_name.name}

Worker Name ID:
{worker_name.id}

Company ID:
{selected_company_id}

Created By:
{request.user.username}
"""
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Worker name added successfully"
        )

        return redirect('worker_name_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker_name/create.html',
        {
            'workers': workers
        }
    )


from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages


@login_required(login_url='login')
def worker_name_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH WORKER NAME
    # =====================================================

    worker_name = get_object_or_404(
        WorkerName,
        pk=pk
    )

    # =====================================================
    # FETCH WORKERS
    # =====================================================

    workers = Worker.objects.filter(
        company_id=selected_company_id
    )

    # =====================================================
    # UPDATE WORKER NAME
    # =====================================================

    if request.method == 'POST':

        # Store old values for activity log
        old_worker = worker_name.worker
        old_name = worker_name.name

        worker_id = request.POST.get('worker')

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not worker_id or not name:

            messages.error(
                request,
                "Worker and name are required"
            )

            return redirect(
                'worker_name_update',
                pk=pk
            )

        # =====================================================
        # FETCH WORKER
        # =====================================================

        worker = get_object_or_404(
            Worker,
            id=worker_id,
            company_id=selected_company_id
        )

        # =====================================================
        # PREVENT DUPLICATE WORKER NAME
        # =====================================================

        duplicate = WorkerName.objects.filter(
            worker=worker,
            name__iexact=name
        ).exclude(
            pk=worker_name.pk
        ).exists()

        if duplicate:

            messages.warning(
                request,
                "Another worker with this name already exists."
            )

            return redirect(
                'worker_name_update',
                pk=pk
            )

        # =====================================================
        # UPDATE WORKER NAME
        # =====================================================

        worker_name.worker = worker

        worker_name.name = name

        worker_name.save()

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Updated Worker Name",

            description=f"""
Worker Name Updated Successfully

Old Worker Team:
{old_worker.name}

New Worker Team:
{worker.name}

Old Worker Name:
{old_name}

New Worker Name:
{worker_name.name}

Worker Name ID:
{worker_name.id}

Company ID:
{selected_company_id}

Updated By:
{request.user.username}
"""
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Worker name updated successfully"
        )

        return redirect('worker_name_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker_name/update.html',
        {
            'worker_name': worker_name,
            'workers': workers,
        }
    )



@login_required(login_url='login')
def worker_name_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    # =====================================================
    # FETCH WORKER NAME
    # =====================================================

    worker_name = get_object_or_404(
        WorkerName,
        pk=pk,
        worker__company_id=selected_company_id
    )

    # Store values before delete
    worker_team = worker_name.worker.name
    worker_person_name = worker_name.name
    worker_name_id = worker_name.id

    # =====================================================
    # DELETE WORKER NAME
    # =====================================================

    if request.method == 'POST':

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Deleted Worker Name",

            description=f"""
Worker Name Deleted Successfully

Worker Team:
{worker_team}

Worker Name:
{worker_person_name}

Worker Name ID:
{worker_name_id}

Company ID:
{selected_company_id}

Deleted By:
{request.user.username}
"""
        )

        # =====================================================
        # DELETE WORKER NAME
        # =====================================================

        worker_name.delete()

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Worker name deleted successfully"
        )

        return redirect('worker_name_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'worker_name/delete.html',
        {
            'worker_name': worker_name
        }
    )

from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse
from django.views.decorators.http import require_POST
@require_POST
def toggle_worker_name_status(request, pk):

    wn = get_object_or_404(
        WorkerName,
        pk=pk,
        worker__company_id=request.session.get('selected_company_id')
    )

    wn.is_active = not wn.is_active
    wn.save(update_fields=['is_active'])

    return JsonResponse({
        'success': True,
        'is_active': wn.is_active
    })



from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from .models import ExpenseCategory

# 📄 INDEX
def expense_category_index(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    ).order_by('name')

    return render(request, 'expense_category/index.html', {
        'categories': categories
    })



# ➕ CREATE
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import ExpenseCategory


# ➕ CREATE
# def expense_category_create(request):

#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         return redirect('dashboard')

#     if request.method == 'POST':
#         name = request.POST.get('name', '').strip()

#         if not name:
#             messages.error(request, "Category name is required")
#             return redirect('expense_category_create')

#         # 🚫 Prevent duplicate category
#         if ExpenseCategory.objects.filter(
#             company_id=selected_company_id,
#             name__iexact=name
#         ).exists():
#             messages.error(request, "Category already exists")
#             return redirect('expense_category_create')

#         ExpenseCategory.objects.create(
#             company_id=selected_company_id,
#             name=name
#         )

#         messages.success(request, "Category created successfully")
#         return redirect('expense_category_index')

#     return render(request, 'expense_category/create.html')




@login_required(login_url='login')
def expense_category_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # CREATE EXPENSE CATEGORY
    # =====================================================

    if request.method == 'POST':

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Category name is required"
            )

            return redirect('expense_category_create')

        # =====================================================
        # PREVENT DUPLICATE CATEGORY
        # =====================================================

        if ExpenseCategory.objects.filter(
            company_id=selected_company_id,
            name__iexact=name
        ).exists():

            messages.error(
                request,
                "Category already exists"
            )

            return redirect('expense_category_create')

        # =====================================================
        # CREATE CATEGORY
        # =====================================================

        category = ExpenseCategory.objects.create(

            company_id=selected_company_id,

            name=name
        )

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Created Expense Category",

            description=f"""
Expense Category Created Successfully

Category Name:
{category.name}

Category ID:
{category.id}

Company:
{company.name}

Created By:
{request.user.username}
"""
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Category created successfully"
        )

        return redirect('expense_category_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'expense_category/create.html'
    )






# ✏️ UPDATE

# def expense_category_update(request, pk):

#     selected_company_id = request.session.get('selected_company_id')

#     if not selected_company_id:
#         return redirect('dashboard')

#     category = get_object_or_404(
#         ExpenseCategory,
#         pk=pk,
#         company_id=selected_company_id
#     )

#     if request.method == 'POST':
#         name = request.POST.get('name', '').strip()

#         if not name:
#             messages.error(request, "Category name is required")
#             return redirect('expense_category_update', pk=pk)

#         # 🚫 Prevent duplicate (exclude current record)
#         if ExpenseCategory.objects.filter(
#             company_id=selected_company_id,
#             name__iexact=name
#         ).exclude(pk=category.pk).exists():

#             messages.error(request, "Another category with this name already exists")
#             return redirect('expense_category_update', pk=pk)

#         category.name = name
#         category.save()

#         messages.success(request, "Category updated successfully")
#         return redirect('expense_category_index')

#     return render(request, 'expense_category/update.html', {
#         'category': category
#     })




@login_required(login_url='login')
def expense_category_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # FETCH CATEGORY
    # =====================================================

    category = get_object_or_404(
        ExpenseCategory,
        pk=pk,
        company_id=selected_company_id
    )

    # =====================================================
    # UPDATE CATEGORY
    # =====================================================

    if request.method == 'POST':

        # Store old value for activity log
        old_name = category.name

        name = request.POST.get(
            'name',
            ''
        ).strip()

        # =====================================================
        # VALIDATION
        # =====================================================

        if not name:

            messages.error(
                request,
                "Category name is required"
            )

            return redirect(
                'expense_category_update',
                pk=pk
            )

        # =====================================================
        # PREVENT DUPLICATE CATEGORY
        # =====================================================

        if ExpenseCategory.objects.filter(
            company_id=selected_company_id,
            name__iexact=name
        ).exclude(
            pk=category.pk
        ).exists():

            messages.error(
                request,
                "Another category with this name already exists"
            )

            return redirect(
                'expense_category_update',
                pk=pk
            )

        # =====================================================
        # UPDATE CATEGORY
        # =====================================================

        category.name = name

        category.save()

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Updated Expense Category",

            description=f"""
Expense Category Updated Successfully

Old Category Name:
{old_name}

New Category Name:
{category.name}

Category ID:
{category.id}

Company:
{company.name}

Updated By:
{request.user.username}
"""
        )

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Category updated successfully"
        )

        return redirect('expense_category_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'expense_category/update.html',
        {
            'category': category
        }
    )




# 🗑 DELETE
@login_required(login_url='login')
def expense_category_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # FETCH CATEGORY
    # =====================================================

    category = get_object_or_404(
        ExpenseCategory,
        pk=pk,
        company_id=selected_company_id
    )

    # Store values before delete
    category_name = category.name
    category_id = category.id

    # =====================================================
    # DELETE CATEGORY
    # =====================================================

    if request.method == 'POST':

        # =====================================================
        # ACTIVITY LOG
        # =====================================================

        ActivityLog.objects.create(

            action="Deleted Expense Category",

            description=f"""
Expense Category Deleted Successfully

Category Name:
{category_name}

Category ID:
{category_id}

Company:
{company.name}

Deleted By:
{request.user.username}
"""
        )

        # =====================================================
        # DELETE CATEGORY
        # =====================================================

        category.delete()

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Category deleted successfully"
        )

        return redirect('expense_category_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'expense_category/delete.html',
        {
            'category': category
        }
    )



# 📄 INDEX
def expense_subcategory_index(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    category_id = request.GET.get('category')

    # ✅ COMPANY FILTER
    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    subcategories = ExpenseSubCategory.objects.select_related('category').filter(
        company_id=selected_company_id
    )

    if category_id:
        subcategories = subcategories.filter(category_id=category_id)

    return render(request, 'expense_subcategory/index.html', {
        'categories': categories,
        'subcategories': subcategories,
        'selected_category': category_id,
    })



# ➕ CREATE
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from .models import ExpenseCategory, ExpenseSubCategory


def expense_subcategory_create(request):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    if request.method == 'POST':
        category_id = request.POST.get('category')
        name = request.POST.get('name', '').strip()

        if not category_id or not name:
            messages.error(request, "Category and Sub-category name are required")
            return redirect('expense_subcategory_create')

        # 🚫 Prevent duplicate subcategory
        if ExpenseSubCategory.objects.filter(
            company_id=selected_company_id,
            category_id=category_id,
            name__iexact=name
        ).exists():

            messages.error(request, "Sub-category already exists for this category")
            return redirect('expense_subcategory_create')

        ExpenseSubCategory.objects.create(
            company_id=selected_company_id,
            category_id=category_id,
            name=name
        )

        messages.success(request, "Sub-category added successfully")
        return redirect('expense_subcategory_index')

    return render(request, 'expense_subcategory/create.html', {
        'categories': categories
    })



# ✏️ UPDATE
def expense_subcategory_update(request, pk):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    subcategory = get_object_or_404(
        ExpenseSubCategory,
        pk=pk,
        company_id=selected_company_id
    )

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    if request.method == 'POST':
        category_id = request.POST.get('category')
        name = request.POST.get('name', '').strip()

        if not category_id or not name:
            messages.error(request, "All fields are required")
            return redirect('expense_subcategory_update', pk=pk)

        # 🚫 Prevent duplicate (exclude current record)
        if ExpenseSubCategory.objects.filter(
            company_id=selected_company_id,
            category_id=category_id,
            name__iexact=name
        ).exclude(pk=subcategory.pk).exists():

            messages.error(request, "Another sub-category with this name already exists")
            return redirect('expense_subcategory_update', pk=pk)

        subcategory.category_id = category_id
        subcategory.name = name
        subcategory.save()

        messages.success(request, "Sub-category updated successfully")
        return redirect('expense_subcategory_index')

    return render(request, 'expense_subcategory/update.html', {
        'subcategory': subcategory,
        'categories': categories
    })



# 🗑 DELETE
def expense_subcategory_delete(request, pk):
    subcategory = get_object_or_404(ExpenseSubCategory, pk=pk)

    if request.method == 'POST':
        subcategory.delete()
        messages.success(request, "Sub-category deleted")
        return redirect('expense_subcategory_index')

    return render(request, 'expense_subcategory/delete.html', {
        'subcategory': subcategory
    })



#expense views will be added here

from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date


def expense_index(request):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    # =========================
    # 📊 BASE QUERYSET
    # =========================
    expenses = Expense.objects.select_related(
        'client',
        'bank',
        'category',
        'subcategory',
        'salary_to',
        'worker_name'
    ).filter(
        client__company_id=selected_company_id
    )

    # =========================
    # 🔍 FILTER VALUES
    # =========================
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_id = request.GET.get('client')
    category_id = request.GET.get('category')
    subcategory_id = request.GET.get('subcategory')
    spend_mode = request.GET.get('spend_mode')
    bank_id = request.GET.get('bank')
    worker_id = request.GET.get('worker')
    worker_name_id = request.GET.get('worker_name')

    # =========================
    # 📅 DATE FILTERS
    # =========================
    if start_date:
        expenses = expenses.filter(
            expense_date__gte=parse_date(start_date)
        )

    if end_date:
        expenses = expenses.filter(
            expense_date__lte=parse_date(end_date)
        )

    # =========================
    # 👤 CLIENT FILTER
    # =========================
    if client_id:
        expenses = expenses.filter(client_id=client_id)

    # =========================
    # 🏷 CATEGORY FILTER
    # =========================
    if category_id:
        expenses = expenses.filter(category_id=category_id)

    # =========================
    # 🏷 SUBCATEGORY FILTER
    # =========================
    if subcategory_id:
        expenses = expenses.filter(subcategory_id=subcategory_id)

    # =========================
    # 💳 SPEND MODE FILTER
    # =========================
    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    # =========================
    # 🏦 BANK FILTER
    # =========================
    if bank_id:
        expenses = expenses.filter(bank_id=bank_id, client__company_id=selected_company_id)

    # =========================
    # 👷 SALARY TEAM FILTER
    # =========================
    if worker_id:
        expenses = expenses.filter(salary_to_id=worker_id)

    # =========================
    # 👤 SALARY WORKER NAME FILTER
    # =========================
    if worker_name_id:
        expenses = expenses.filter(worker_name_id=worker_name_id)

    # =========================
    # 📅 ORDERING + LIMIT 100 per page
    # =========================
    expenses = expenses.order_by(
        '-expense_date',
        '-id'
    )[:100]

    # =========================
    # 📋 DROPDOWN DATA
    # =========================
    clients = Client.objects.filter(
        company_id=selected_company_id
    )

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    subcategories = ExpenseSubCategory.objects.filter(
        category__company_id=selected_company_id
    )

    workers = Worker.objects.filter(
        company_id=selected_company_id
    )

    worker_names = WorkerName.objects.filter(
        worker__company_id=selected_company_id
    ).select_related('worker')

    cash_bank = Bank.objects.filter(
        name__iexact='cash'
    ).first()

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    ).exclude(
        name__iexact='cash'
    )

    # =========================
    # 📤 RENDER
    # =========================
    return render(request, 'expense/index.html', {
        'expenses': expenses,

        # Selected values (for form state persistence)
        'start_date': start_date,
        'end_date': end_date,
        'selected_client': client_id,
        'selected_category': category_id,
        'selected_subcategory': subcategory_id,
        'selected_spend_mode': spend_mode,
        'selected_bank': bank_id,
        'selected_worker': worker_id,
        'selected_worker_name': worker_name_id,

        # Dropdowns
        'clients': clients,
        'categories': categories,
        'subcategories': subcategories,
        'workers': workers,
        'worker_names': worker_names,
        'cash_bank': cash_bank,
        'banks': banks,
    })




from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from django.http import HttpResponse
from django.shortcuts import redirect
from django.utils.dateparse import parse_date
from decimal import Decimal


def expense_pdf_export(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    expenses = Expense.objects.select_related(
        'client',
        'bank',
        'category',
        'subcategory',
        'salary_to',
        'worker_name'
    ).filter(
        client__company_id=selected_company_id
    )

    # =============================
    # FILTERS
    # =============================
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_id = request.GET.get('client')
    category_id = request.GET.get('category')
    spend_mode = request.GET.get('spend_mode')
    bank_id = request.GET.get('bank')
    worker_id = request.GET.get('worker')

    if start_date:
        expenses = expenses.filter(expense_date__gte=parse_date(start_date))

    if end_date:
        expenses = expenses.filter(expense_date__lte=parse_date(end_date))

    if client_id:
        expenses = expenses.filter(client_id=client_id)

    if category_id:
        expenses = expenses.filter(category_id=category_id)

    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    if bank_id:
        expenses = expenses.filter(bank_id=bank_id)

    if worker_id:
        expenses = expenses.filter(salary_to_id=worker_id)

    expenses = expenses.order_by('expense_date')

    # =============================
    # PDF SETUP
    # =============================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="expenses_report.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()

    # 🔥 Custom small text style
    small_style = ParagraphStyle(
        name='Small',
        fontSize=8,
        leading=10
    )

    bold_style = ParagraphStyle(
        name='BoldSmall',
        fontSize=9,
        leading=11,
        fontName='Helvetica-Bold'
    )

    elements = []

    company = Company.objects.get(id=selected_company_id)

    elements.append(Paragraph(
        f"<b>Expense Report – {company.name}</b>",
        styles['Title']
    ))

    if start_date or end_date:
        elements.append(
            Paragraph(
                f"Period: {start_date or '—'} to {end_date or '—'}",
                styles['Italic']
            )
        )

    elements.append(Spacer(1, 12))

    # =============================
    # TABLE DATA
    # =============================
    table_data = [[
        Paragraph('Date', bold_style),
        Paragraph('Client', bold_style),
        Paragraph('Category', bold_style),
        Paragraph('Sub-Category', bold_style),
        Paragraph('Description', bold_style),
        Paragraph('Worker Team', bold_style),
        Paragraph('Worker Name', bold_style),
        Paragraph('Mode', bold_style),
        Paragraph('Bank', bold_style),
        Paragraph('Amount', bold_style),
    ]]

    total_amount = Decimal('0.00')

    for e in expenses:
        total_amount += e.amount

        table_data.append([
            Paragraph(e.expense_date.strftime('%d-%m-%Y'), small_style),
            Paragraph(e.client.name, small_style),
            Paragraph(e.category.name if e.category else '—', small_style),
            Paragraph(e.subcategory.name if e.subcategory else '—', small_style),
            Paragraph(e.description if e.description else '—', small_style),
            Paragraph(e.salary_to.name if e.salary_to else '—', small_style),
            Paragraph(e.worker_name.name if e.worker_name else '—', small_style),
            Paragraph(e.spend_mode.capitalize(), small_style),
            Paragraph(e.bank.name if e.bank else 'Cash', small_style),
            Paragraph(f"{e.amount:.2f}", small_style),
        ])

    # 🔥 Adjusted widths (important)
    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[60, 90, 80, 90, 140, 90, 90, 50, 80, 70]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),

        # 🔥 Wrap + alignment fixes
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (-1,1), (-1,-1), 'RIGHT'),

        # Padding fix
        ('LEFTPADDING', (0,0), (-1,-1), 5),
        ('RIGHTPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # =============================
    # TOTAL
    # =============================
    total_table = Table(
        [[
            Paragraph('TOTAL EXPENSE', bold_style),
            Paragraph(f"Rs. {total_amount:.2f}", bold_style)
        ]],
        colWidths=[400, 150]
    )

    total_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,-1), colors.whitesmoke),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 8),
    ]))

    elements.append(total_table)

    doc.build(elements)
    return response



#export as excel will be added here


def expense_excel_export(request):

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    expenses = Expense.objects.select_related(
        'client',
        'bank',
        'category',
        'subcategory',
        'salary_to',
        'worker_name'
    ).filter(
        client__company_id=selected_company_id
    )

    # =============================
    # FILTERS
    # =============================

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    client_id = request.GET.get('client')
    category_id = request.GET.get('category')
    spend_mode = request.GET.get('spend_mode')
    bank_id = request.GET.get('bank')
    worker_id = request.GET.get('worker')

    if start_date:
        expenses = expenses.filter(
            expense_date__gte=parse_date(start_date)
        )

    if end_date:
        expenses = expenses.filter(
            expense_date__lte=parse_date(end_date)
        )

    if client_id:
        expenses = expenses.filter(client_id=client_id)

    if category_id:
        expenses = expenses.filter(category_id=category_id)

    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    if bank_id:
        expenses = expenses.filter(bank_id=bank_id)

    if worker_id:
        expenses = expenses.filter(salary_to_id=worker_id)

    expenses = expenses.order_by('expense_date')

    # =============================
    # WORKBOOK
    # =============================

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # =============================
    # STYLES
    # =============================

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    total_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    thin = Side(style='thin', color='CCCCCC')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =============================
    # COMPANY
    # =============================

    company = Company.objects.get(id=selected_company_id)

    # =============================
    # TITLE
    # =============================

    ws.merge_cells('A1:K1')

    title = ws['A1']

    title.value = f"EXPENSE REPORT – {company.name}"

    title.font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    title.fill = header_fill
    title.alignment = center

    # =============================
    # PERIOD
    # =============================

    if start_date or end_date:

        ws.merge_cells('A2:K2')

        period = ws['A2']

        period.value = (
            f"Period: {start_date or '—'} to {end_date or '—'}"
        )

        period.font = Font(
            italic=True
        )

    # =============================
    # HEADERS
    # =============================

    headers = [
        'Date',
        'Client',
        'Location',
        'Category',
        'Sub-Category',
        'Description',
        'Worker Team',
        'Worker Name',
        'Mode',
        'Bank',
        'Amount',
    ]

    row = 4

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row, col_num)

        cell.value = header
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row += 1

    # =============================
    # DATA
    # =============================

    total_amount = Decimal('0.00')

    for e in expenses:

        total_amount += e.amount

        data = [
            e.expense_date.strftime('%d-%m-%Y'),
            e.client.name,
            e.client.location if e.client.location else '—',
            e.category.name if e.category else '—',
            e.subcategory.name if e.subcategory else '—',
            e.description if e.description else '—',
            e.salary_to.name if e.salary_to else '—',
            e.worker_name.name if e.worker_name else '—',
            e.spend_mode.capitalize(),
            e.bank.name if e.bank else 'Cash',
            float(e.amount),
        ]

        for col_num, value in enumerate(data, 1):

            cell = ws.cell(row, col_num)

            cell.value = value
            cell.border = border

            if col_num == 11:
                cell.alignment = right
            else:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        row += 1

    # =============================
    # TOTAL ROW
    # =============================

    totals = [
        'TOTAL EXPENSE',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        float(total_amount)
    ]

    for col_num, value in enumerate(totals, 1):

        cell = ws.cell(row, col_num)

        cell.value = value
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border

        if col_num == 11:
            cell.alignment = right

    # =============================
    # COLUMN WIDTHS
    # =============================

    widths = {
        1: 15,   # Date
        2: 25,   # Client
        3: 25,   # Location
        4: 20,   # Category
        5: 22,   # Sub Category
        6: 40,   # Description
        7: 22,   # Worker Team
        8: 22,   # Worker Name
        9: 14,   # Mode
        10: 20,  # Bank
        11: 18,  # Amount
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # =============================
    # RESPONSE
    # =============================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename="expenses_report.xlsx"'
    )

    wb.save(response)

    return response



from django.http import JsonResponse
from decimal import Decimal

def check_expense_duplicate(request):

    selected_company_id = request.session.get('selected_company_id')

    client_id = request.GET.get('client')
    category_id = request.GET.get('category')
    subcategory_id = request.GET.get('subcategory')
    expense_date = request.GET.get('expense_date')
    amount = request.GET.get('amount')

    if not (client_id and category_id and expense_date and amount):
        return JsonResponse({"duplicate": False})

    try:
        amount = Decimal(str(amount).replace(',', '').strip())
    except:
        return JsonResponse({"duplicate": False})

    duplicate = Expense.objects.filter(
        client_id=client_id,
        client__company_id=selected_company_id,
        category_id=category_id,
        subcategory_id=subcategory_id,
        expense_date=expense_date,
        amount=amount
    ).first()

    if duplicate:
        return JsonResponse({
            "duplicate": True,
            "client": duplicate.client.name,
            "amount": str(duplicate.amount),
            "date": str(duplicate.expense_date)
        })

    return JsonResponse({"duplicate": False})




from decimal import Decimal
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect
from django.db.models import Q

from decimal import Decimal
from django.db import transaction
from django.contrib import messages
from django.shortcuts import redirect, render



@login_required(login_url='login')
def expense_create(request):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =========================
    # DROPDOWNS
    # =========================

    clients = Client.objects.filter(
        company_id=selected_company_id,
        is_active=True
    )

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    subcategories = ExpenseSubCategory.objects.filter(
        category__company_id=selected_company_id
    )

    workers = Worker.objects.filter(
        company_id=selected_company_id
    )

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    ).order_by('name')

    cash_bank = Bank.objects.filter(
        company_id=selected_company_id,
        name__iexact='cash'
    ).first()

    worker_names = WorkerName.objects.filter(
        worker__company_id=selected_company_id,
        is_active=True
    ).select_related('worker')

    # =========================
    # POST
    # =========================

    if request.method == 'POST':

        client_id = request.POST.get('client')

        bank_id = request.POST.get('bank')

        category_id = request.POST.get('category')

        subcategory_id = request.POST.get('subcategory')

        salary_to_id = request.POST.get('salary_to')

        worker_name_id = request.POST.get('worker_name')

        description = request.POST.get(
            'description',
            ''
        ).strip()

        spend_mode = request.POST.get('spend_mode')

        expense_date = request.POST.get('expense_date')

        amount_raw = request.POST.get('amount')

        # =========================
        # VALIDATE AMOUNT
        # =========================

        if not amount_raw:

            messages.error(
                request,
                "Amount is required"
            )

            return redirect('expense_create')

        try:

            amount = Decimal(
                str(amount_raw)
                .replace(',', '')
                .strip()
            )

        except:

            messages.error(
                request,
                "Invalid amount"
            )

            return redirect('expense_create')

        # =========================
        # VALIDATIONS
        # =========================

        client = clients.filter(
            id=client_id
        ).first()

        if not client:

            messages.error(
                request,
                "Invalid client"
            )

            return redirect('expense_create')

        category = (
            categories.filter(id=category_id).first()
            if category_id else None
        )

        subcategory = None

        if subcategory_id:

            subcategory = subcategories.filter(
                id=subcategory_id,
                category_id=category_id
            ).first()

            if not subcategory:

                messages.error(
                    request,
                    "Invalid sub-category selection"
                )

                return redirect('expense_create')

        worker = (
            workers.filter(id=salary_to_id).first()
            if salary_to_id else None
        )

        worker_name = None

        if worker_name_id:

            worker_name = worker_names.filter(
                id=worker_name_id
            ).first()

        # =========================
        # DUPLICATE DETECTION
        # =========================

        duplicate_expense = Expense.objects.filter(
            client_id=client_id,
            client__company_id=selected_company_id,
            category_id=category_id,
            subcategory_id=subcategory_id,
            expense_date=expense_date,
            amount=amount
        ).first()

        duplicate_warning = False

        if duplicate_expense:
            duplicate_warning = True

        # =========================
        # SAVE EXPENSE
        # =========================

        with transaction.atomic():

            # =========================
            # CASH MODE
            # =========================

            if spend_mode == Expense.CASH:

                bank = cash_bank

            # =========================
            # BANK MODE
            # =========================

            else:

                bank = banks.filter(
                    id=bank_id
                ).first()

                if not bank:

                    messages.error(
                        request,
                        "Invalid bank"
                    )

                    return redirect('expense_create')

            # =========================
            # CREATE EXPENSE
            # =========================

            expense = Expense.objects.create(

                client=client,

                bank=bank,

                category=category,

                subcategory=subcategory,

                salary_to=worker,

                worker_name_id=worker_name_id,

                description=description,

                amount=amount,

                spend_mode=spend_mode,

                expense_date=expense_date
            )

            # =========================
            # RECALCULATE BANK
            # =========================

            recalculate_banks(bank)

            bank.refresh_from_db()

            # =========================
            # ACTIVITY LOG
            # =========================

            log_activity(

    request=request,

    action="Created Expense",

    action_type='Created',

    description=f"""
Expense Created Successfully

Expense ID:
{expense.id}

Client:
{client.name}

Location:
{client.location or 'N/A'}

Category:
{category.name if category else 'N/A'}

Sub Category:
{subcategory.name if subcategory else 'N/A'}

Worker Team:
{worker.name if worker else 'N/A'}

Worker Name:
{worker_name.name if worker_name else 'N/A'}

Description:
{description or 'N/A'}

Expense Amount:
Rs. {amount}

Spend Mode:
{spend_mode}

Expense Date:
{expense_date}

Bank:
{bank.name if bank else 'Cash'}

Updated Bank Balance:
Rs. {bank.available_balance}

Company:
{company.name}
"""
)

        # =========================
        # DUPLICATE WARNING
        # =========================

        if duplicate_warning:

            messages.warning(
                request,
                "⚠ A similar expense already existed. This entry was saved anyway."
            )

        # =========================
        # SUCCESS MESSAGE
        # =========================

        messages.success(
            request,
            "Expense added successfully"
        )

        return redirect('expense_index')

    # =========================
    # GET PAGE
    # =========================

    return render(
        request,
        'expense/create.html',
        {
            'clients': clients,
            'banks': banks,
            'cash_bank': cash_bank,
            'categories': categories,
            'subcategories': subcategories,
            'workers': workers,
            'selected_company_id': selected_company_id,
            'worker_names': worker_names,
        }
    )




from django.shortcuts import get_object_or_404, redirect, render
from django.db import transaction
from django.contrib import messages
from decimal import Decimal





@login_required(login_url='login')
def expense_update(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # FETCH EXPENSE
    # =====================================================

    expense = get_object_or_404(
        Expense,
        pk=pk,
        client__company_id=selected_company_id
    )

    # Store old values for activity log
    old_client = expense.client
    old_category = expense.category
    old_subcategory = expense.subcategory
    old_worker = expense.salary_to
    old_worker_name = expense.worker_name
    old_bank = expense.bank
    old_amount = expense.amount
    old_spend_mode = expense.spend_mode
    old_description = expense.description
    old_expense_date = expense.expense_date

    # =========================
    # DROPDOWNS
    # =========================

    clients = Client.objects.filter(
        company_id=selected_company_id,
        is_active=True
    )

    categories = ExpenseCategory.objects.filter(
        company_id=selected_company_id
    )

    subcategories = ExpenseSubCategory.objects.filter(
        category__company_id=selected_company_id
    )

    workers = Worker.objects.filter(
        company_id=selected_company_id
    )

    worker_names = WorkerName.objects.filter(
        worker__company_id=selected_company_id,
        is_active=True
    ).select_related('worker')

    banks = Bank.objects.filter(
        company_id=selected_company_id,
        is_active=True
    ).order_by('name')

    cash_bank = Bank.objects.filter(
        company_id=selected_company_id,
        name__iexact='cash'
    ).first()

    # =========================
    # POST
    # =========================

    if request.method == "POST":

        client_id = request.POST.get('client')

        category_id = request.POST.get('category')

        subcategory_id = request.POST.get('subcategory')

        salary_to_id = request.POST.get('salary_to')

        worker_name_id = request.POST.get('worker_name')

        bank_id = request.POST.get('bank')

        description = request.POST.get(
            'description',
            ''
        ).strip()

        spend_mode = request.POST.get('spend_mode')

        expense_date = request.POST.get('expense_date')

        # =========================
        # VALIDATE AMOUNT
        # =========================

        try:

            new_amount = Decimal(
                str(request.POST.get('amount'))
                .replace(',', '')
                .strip()
            )

        except:

            messages.error(
                request,
                "Invalid amount"
            )

            return redirect(
                'expense_update',
                pk=pk
            )

        # =========================
        # VALIDATIONS
        # =========================

        client = clients.filter(
            id=client_id
        ).first()

        if not client:

            messages.error(
                request,
                "Invalid client"
            )

            return redirect(
                'expense_update',
                pk=pk
            )

        category = (
            categories.filter(id=category_id).first()
            if category_id else None
        )

        subcategory = None

        if subcategory_id:

            subcategory = subcategories.filter(
                id=subcategory_id,
                category_id=category_id
            ).first()

        worker = None
        worker_name = None

        # =========================
        # SALARY CATEGORY LOGIC
        # =========================

        if category and category.name.lower() == "salary":

            if salary_to_id:

                worker = workers.filter(
                    id=salary_to_id
                ).first()

            if worker_name_id:

                worker_name = worker_names.filter(
                    id=worker_name_id,
                    worker=worker
                ).first()

        # =========================
        # DUPLICATE CHECK
        # =========================

        duplicate_expense = Expense.objects.filter(
            client_id=client_id,
            client__company_id=selected_company_id,
            category_id=category_id,
            subcategory_id=subcategory_id,
            expense_date=expense_date,
            amount=new_amount
        ).exclude(
            pk=expense.pk
        ).first()

        duplicate_warning = False

        if duplicate_expense:
            duplicate_warning = True

        # =========================
        # UPDATE EXPENSE
        # =========================

        with transaction.atomic():

            # =========================
            # RECALCULATE OLD BANK
            # =========================

            # if old_bank:
            #     old_bank.recalculate_balance()

            # =========================
            # CASH MODE
            # =========================

            if spend_mode == Expense.CASH:

                bank = cash_bank

            # =========================
            # BANK MODE
            # =========================

            else:

                bank = banks.filter(
                    id=bank_id
                ).first()

                if not bank:

                    messages.error(
                        request,
                        "Invalid bank"
                    )

                    return redirect(
                        'expense_update',
                        pk=pk
                    )

            # =========================
            # SAVE EXPENSE
            # =========================

            expense.client = client
            expense.category = category
            expense.subcategory = subcategory
            expense.salary_to = worker
            expense.worker_name = worker_name
            expense.bank = bank
            expense.description = description
            expense.amount = new_amount
            expense.spend_mode = spend_mode
            expense.expense_date = expense_date

            expense.save()

            # =========================
            # RECALCULATE OLD BANK
            # =========================

            if old_bank:

                recalculate_banks(bank)

                bank.refresh_from_db()

            # =========================
            # RECALCULATE NEW BANK
            # =========================

            if bank:

                bank.refresh_from_db()

                bank.recalculate_balance()

                bank.refresh_from_db()

            # =========================
            # ACTIVITY LOG
            # =========================

            log_activity(

        request=request,

        action="Updated Expense",

        action_type='Updated',

        description=f"""
    Expense Updated Successfully

    Expense ID:
    {expense.id}

    --------------------------------------------------

    OLD VALUES

    Client:
    {old_client.name}

    Location:
    {old_client.location or 'N/A'}

    Category:
    {old_category.name if old_category else 'N/A'}

    Sub Category:
    {old_subcategory.name if old_subcategory else 'N/A'}

    Worker Team:
    {old_worker.name if old_worker else 'N/A'}

    Worker Name:
    {old_worker_name.name if old_worker_name else 'N/A'}

    Description:
    {old_description or 'N/A'}

    Amount:
    Rs. {old_amount}

    Spend Mode:
    {old_spend_mode}

    Expense Date:
    {old_expense_date}

    Bank:
    {old_bank.name if old_bank else 'N/A'}

    --------------------------------------------------

    NEW VALUES

    Client:
    {client.name}

    Location:
    {client.location or 'N/A'}

    Category:
    {category.name if category else 'N/A'}

    Sub Category:
    {subcategory.name if subcategory else 'N/A'}

    Worker Team:
    {worker.name if worker else 'N/A'}

    Worker Name:
    {worker_name.name if worker_name else 'N/A'}

    Description:
    {description or 'N/A'}

    Amount:
    Rs. {new_amount}

    Spend Mode:
    {spend_mode}

    Expense Date:
    {expense_date}

    Bank:
    {bank.name if bank else 'Cash'}

    Updated Bank Balance:
    Rs. {bank.available_balance}

    --------------------------------------------------

    Company:
    {company.name}
    """
    )

        # =========================
        # DUPLICATE WARNING
        # =========================

        if duplicate_warning:

            messages.warning(
                request,
                "⚠ Similar expense already exists. Update saved anyway."
            )

        # =========================
        # SUCCESS MESSAGE
        # =========================

        messages.success(
            request,
            "Expense updated successfully"
        )

        return redirect('expense_index')

    # =========================
    # GET PAGE
    # =========================

    return render(
        request,
        "expense/update.html",
        {
            "expense": expense,
            "clients": clients,
            "banks": banks,
            "cash_bank": cash_bank,
            "categories": categories,
            "subcategories": subcategories,
            "workers": workers,
            "worker_names": worker_names,
            "selected_company_id": selected_company_id,
        }
    )




from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.db import transaction


@login_required(login_url='login')
def expense_delete(request, pk):

    # =====================================================
    # COMPANY FROM SESSION
    # =====================================================

    selected_company_id = request.session.get(
        'selected_company_id'
    )

    if not selected_company_id:
        return redirect('dashboard')

    company = get_object_or_404(
        Company,
        id=selected_company_id
    )

    # =====================================================
    # FETCH EXPENSE
    # =====================================================

    expense = get_object_or_404(
        Expense,
        pk=pk,
        client__company_id=selected_company_id
    )

    bank = expense.bank

    # Store values before delete
    client_name = expense.client.name
    client_location = expense.client.location

    category_name = (
        expense.category.name
        if expense.category else 'N/A'
    )

    subcategory_name = (
        expense.subcategory.name
        if expense.subcategory else 'N/A'
    )

    worker_team = (
        expense.salary_to.name
        if expense.salary_to else 'N/A'
    )

    worker_person = (
        expense.worker_name.name
        if expense.worker_name else 'N/A'
    )

    description = expense.description
    amount = expense.amount
    spend_mode = expense.spend_mode
    expense_date = expense.expense_date

    bank_name = (
        bank.name if bank else 'Cash'
    )

    expense_id = expense.id

    # =====================================================
    # DELETE EXPENSE
    # =====================================================

    if request.method == 'POST':

        with transaction.atomic():

            # Store old bank balance
            old_balance = (
                bank.available_balance
                if bank else Decimal('0.00')
            )

            # =====================================================
            # ACTIVITY LOG
            # =====================================================

            log_activity(

            request=request,

            action="Deleted Expense",

            action_type='Deleted',

            description=f"""
        Expense Deleted Successfully

        Expense ID:
        {expense_id}

        Client:
        {client_name}

        Location:
        {client_location or 'N/A'}

        Category:
        {category_name}

        Sub Category:
        {subcategory_name}

        Worker Team:
        {worker_team}

        Worker Name:
        {worker_person}

        Description:
        {description or 'N/A'}

        Expense Amount:
        Rs. {amount}

        Spend Mode:
        {spend_mode}

        Expense Date:
        {expense_date}

        Bank:
        {bank_name}

        Old Bank Balance:
        Rs. {old_balance}

        Company:
        {company.name}
        """
        )
            # =====================================================
            # DELETE EXPENSE
            # =====================================================

            expense.delete()

            # =====================================================
            # RECALCULATE BANK
            # =====================================================

            if bank:

                recalculate_banks(bank)

                bank.refresh_from_db()

                # =====================================================
                # UPDATE LOG WITH NEW BALANCE
                # =====================================================

                latest_log = ActivityLog.objects.latest('id')

                latest_log.description += f"""

New Bank Balance:
Rs. {bank.available_balance}
"""

                latest_log.save()

        # =====================================================
        # SUCCESS MESSAGE
        # =====================================================

        messages.success(
            request,
            "Expense deleted successfully"
        )

        return redirect('expense_index')

    # =====================================================
    # GET REQUEST
    # =====================================================

    return render(
        request,
        'expense/delete.html',
        {
            'expense': expense
        }
    )






from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404

@require_POST
def toggle_client_status(request, pk):

    company_id = request.session.get('selected_company_id')

    if not company_id:
        return JsonResponse({'success': False, 'error': 'No company selected'}, status=400)

    client = get_object_or_404(
        Client,
        pk=pk,
        company_id=company_id
    )

    # 🔥 TOGGLE
    client.is_active = not bool(client.is_active)
    client.save(update_fields=['is_active'])

    # 🔥 DEBUG (optional)
    print(f"[TOGGLE] Client {client.id} → {client.is_active}")

    return JsonResponse({
        'success': True,
        'is_active': client.is_active  # 🔥 IMPORTANT
    })





from django.shortcuts import render, redirect
from django.utils.dateparse import parse_date
from django.db.models import Sum
from decimal import Decimal


def salary_index(request):

    # 🏢 COMPANY FROM SESSION
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    # =========================
    # 📊 BASE QUERYSET (ONLY SALARY)
    # =========================
    expenses = Expense.objects.select_related(
        'client', 'bank', 'category', 'salary_to', 'worker_name'
    ).filter(
        client__company_id=selected_company_id,
        category__name__iexact='salary'
    )

    # =========================
    # 🔍 FILTER VALUES
    # =========================
    worker_team_id = request.GET.get('worker')          # Team
    worker_name_id = request.GET.get('worker_name')     # Individual
    client_id = request.GET.get('client')
    spend_mode = request.GET.get('spend_mode')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # =========================
    # 👷 TEAM FILTER
    # =========================
    if worker_team_id:
        expenses = expenses.filter(salary_to_id=worker_team_id)

    # =========================
    # 👷‍♂️ WORKER NAME FILTER
    # =========================
    if worker_name_id:
        expenses = expenses.filter(worker_name_id=worker_name_id)

    # =========================
    # 👤 CLIENT FILTER
    # =========================
    if client_id:
        expenses = expenses.filter(client_id=client_id)

    # =========================
    # 💳 SPEND MODE FILTER
    # =========================
    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    # =========================
    # 📅 DATE FILTERS
    # =========================
    if start_date:
        expenses = expenses.filter(expense_date__gte=parse_date(start_date))

    if end_date:
        expenses = expenses.filter(expense_date__lte=parse_date(end_date))

    expenses = expenses.order_by('-expense_date')

    # =========================
    # 💰 TOTAL SALARY
    # =========================
    total_salary = expenses.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    return render(request, 'salary/index.html', {
        'expenses': expenses,
        'total_salary': total_salary,

        # selected filters
        'selected_worker': worker_team_id,
        'selected_worker_name': worker_name_id,
        'selected_client': client_id,
        'selected_spend_mode': spend_mode,
        'start_date': start_date,
        'end_date': end_date,

        # DROPDOWNS
        'workers': Worker.objects.filter(
            company_id=selected_company_id
        ),

        'worker_names': WorkerName.objects.filter(
            worker__company_id=selected_company_id
        ).select_related('worker'),

        'clients': Client.objects.filter(
            company_id=selected_company_id
        ),
    })




from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect
from django.utils.dateparse import parse_date
from django.db.models import Sum
from decimal import Decimal

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


def salary_pdf(request):

    # =========================
    # 🏢 COMPANY CHECK
    # =========================
    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    worker_id = request.GET.get('worker')
    worker_name_id = request.GET.get('worker_name')
    client_id = request.GET.get('client')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    spend_mode = request.GET.get('spend_mode')

    # =========================
    # 📊 BASE QUERYSET
    # =========================
    expenses = Expense.objects.select_related(
        'client', 'bank', 'salary_to', 'worker_name', 'category'
    ).filter(
        client__company_id=selected_company_id,
        category__name__iexact='salary'
    )

    # 👷 Worker Team Filter
    if worker_id:
        expenses = expenses.filter(salary_to_id=worker_id)

    # 👤 Worker Name Filter
    if worker_name_id:
        expenses = expenses.filter(worker_name_id=worker_name_id)

    # 👥 Client Filter
    if client_id:
        expenses = expenses.filter(client_id=client_id)

    # 💳 Spend Mode
    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    # 📅 Date Filters
    if start_date:
        expenses = expenses.filter(expense_date__gte=parse_date(start_date))

    if end_date:
        expenses = expenses.filter(expense_date__lte=parse_date(end_date))

    expenses = expenses.order_by('expense_date')

    total_salary = expenses.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # =========================
    # 📄 PDF RESPONSE
    # =========================
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="salary_statement.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=30,
        bottomMargin=30
    )

    styles = getSampleStyleSheet()
    elements = []

    # =========================
    # 🧾 HEADER
    # =========================
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Title'],
        fontSize=18,
        spaceAfter=6
    )

    elements.append(Paragraph("Salary Statement Report", header_style))
    elements.append(Spacer(1, 6))

    # Dynamic Header Info
    elements.append(Paragraph(
        f"<b>Worker Team:</b> {expenses.first().salary_to.name if expenses and worker_id else 'All'}",
        styles['Normal']
    ))

    elements.append(Paragraph(
        f"<b>Worker Name:</b> {expenses.first().worker_name.name if expenses and worker_name_id else 'All'}",
        styles['Normal']
    ))

    elements.append(Paragraph(
        f"<b>Period:</b> {start_date or '—'} to {end_date or '—'}",
        styles['Normal']
    ))

    elements.append(Spacer(1, 20))

    # =========================
    # 📊 TABLE HEADER
    # =========================
    table_data = [[
        '#', 'Date', 'Client',
        'Team', 'Worker',
        'Mode', 'Description',
        'Bank', 'Amount (Rs)'
    ]]

    # =========================
    # 📄 TABLE ROWS
    # =========================
    for idx, e in enumerate(expenses, start=1):

        table_data.append([
            idx,
            e.expense_date.strftime('%d-%m-%Y'),
            e.client.name if e.client else '—',
            e.salary_to.name if e.salary_to else '—',
            e.worker_name.name if e.worker_name else '—',
            e.spend_mode.capitalize(),
            e.description,
            e.bank.name if e.bank else 'Cash',
            f"{e.amount:,.2f}"
        ])

    table = Table(
        table_data,
        repeatRows=1,
        colWidths=[35, 70, 120, 100, 100, 70, 160, 90, 90]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E9EFF5')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),

        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (-1, 1), (-1, -1), 'RIGHT'),

        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 25))

    # =========================
    # 💰 TOTAL BOX
    # =========================
    total_table = Table(
        [['Total Salary Paid', f"Rs. {total_salary:,.2f}"]],
        colWidths=[500, 150]
    )

    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F5F7FA')),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))

    elements.append(total_table)

    doc.build(elements)
    return response



#export as excel will be added here

def salary_excel_export(request):

    # =========================
    # COMPANY CHECK
    # =========================

    selected_company_id = request.session.get('selected_company_id')

    if not selected_company_id:
        return redirect('dashboard')

    worker_id = request.GET.get('worker')
    worker_name_id = request.GET.get('worker_name')
    client_id = request.GET.get('client')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    spend_mode = request.GET.get('spend_mode')

    # =========================
    # BASE QUERYSET
    # =========================

    expenses = Expense.objects.select_related(
        'client',
        'bank',
        'salary_to',
        'worker_name',
        'category'
    ).filter(
        client__company_id=selected_company_id,
        category__name__iexact='salary'
    )

    # Worker Team Filter
    if worker_id:
        expenses = expenses.filter(salary_to_id=worker_id)

    # Worker Name Filter
    if worker_name_id:
        expenses = expenses.filter(worker_name_id=worker_name_id)

    # Client Filter
    if client_id:
        expenses = expenses.filter(client_id=client_id)

    # Spend Mode
    if spend_mode in ['cash', 'cheque']:
        expenses = expenses.filter(spend_mode=spend_mode)

    # Date Filters
    if start_date:
        expenses = expenses.filter(
            expense_date__gte=parse_date(start_date)
        )

    if end_date:
        expenses = expenses.filter(
            expense_date__lte=parse_date(end_date)
        )

    expenses = expenses.order_by('expense_date')

    total_salary = expenses.aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0.00')

    # =========================
    # WORKBOOK
    # =========================

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Statement"

    # =========================
    # STYLES
    # =========================

    header_fill = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    total_fill = PatternFill(
        start_color="D9D9D9",
        end_color="D9D9D9",
        fill_type="solid"
    )

    white_font = Font(
        color="FFFFFF",
        bold=True
    )

    bold_font = Font(
        bold=True
    )

    center = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    right = Alignment(
        horizontal="right",
        vertical="center"
    )

    thin = Side(style='thin', color='CCCCCC')

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # =========================
    # TITLE
    # =========================

    ws.merge_cells('A1:J1')

    title = ws['A1']

    title.value = "SALARY STATEMENT REPORT"

    title.font = Font(
        bold=True,
        size=18,
        color="FFFFFF"
    )

    title.fill = header_fill
    title.alignment = center

    # =========================
    # FILTER INFO
    # =========================

    ws['A3'] = "Worker Team"
    ws['B3'] = (
        expenses.first().salary_to.name
        if expenses and worker_id else 'All'
    )

    ws['A4'] = "Worker Name"
    ws['B4'] = (
        expenses.first().worker_name.name
        if expenses and worker_name_id else 'All'
    )

    ws['A5'] = "Period"
    ws['B5'] = f"{start_date or '—'} to {end_date or '—'}"

    ws['A3'].font = bold_font
    ws['A4'].font = bold_font
    ws['A5'].font = bold_font

    # =========================
    # HEADERS
    # =========================

    headers = [
        '#',
        'Date',
        'Client',
        'Location',
        'Team',
        'Worker',
        'Mode',
        'Description',
        'Bank',
        'Amount (Rs)'
    ]

    row = 7

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row, col_num)

        cell.value = header
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row += 1

    # =========================
    # DATA
    # =========================

    for idx, e in enumerate(expenses, start=1):

        data = [
            idx,
            e.expense_date.strftime('%d-%m-%Y'),
            e.client.name if e.client else '—',
            e.client.location if e.client and e.client.location else '—',
            e.salary_to.name if e.salary_to else '—',
            e.worker_name.name if e.worker_name else '—',
            e.spend_mode.capitalize(),
            e.description if e.description else '—',
            e.bank.name if e.bank else 'Cash',
            float(e.amount)
        ]

        for col_num, value in enumerate(data, 1):

            cell = ws.cell(row, col_num)

            cell.value = value
            cell.border = border

            if col_num == 10:
                cell.alignment = right
            else:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

        row += 1

    # =========================
    # TOTAL ROW
    # =========================

    totals = [
        'TOTAL SALARY',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        float(total_salary)
    ]

    for col_num, value in enumerate(totals, 1):

        cell = ws.cell(row, col_num)

        cell.value = value
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border

        if col_num == 10:
            cell.alignment = right

    # =========================
    # COLUMN WIDTHS
    # =========================

    widths = {
        1: 8,    # #
        2: 15,   # Date
        3: 25,   # Client
        4: 25,   # Location
        5: 22,   # Team
        6: 22,   # Worker
        7: 14,   # Mode
        8: 40,   # Description
        9: 20,   # Bank
        10: 18,  # Amount
    }

    for col, width in widths.items():

        ws.column_dimensions[
            get_column_letter(col)
        ].width = width

    # =========================
    # RESPONSE
    # =========================

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename="salary_statement.xlsx"'
    )

    wb.save(response)

    return response

    




from .models import AppSettings, BackupHistory
from django.core.mail import send_mail
from django.utils import timezone
from zoneinfo import ZoneInfo
import os
import zipfile
import subprocess
from datetime import timedelta
from django.conf import settings
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect


@login_required
def database_backup(request):
    try:
        now_dt = timezone.now()
        month_folder = now_dt.strftime('%Y-%m')
        timestamp = now_dt.strftime('%Y-%m-%d_%H-%M')

        backup_root = os.path.join('/opt/backups/eliteAcc', month_folder)
        os.makedirs(backup_root, exist_ok=True)

        sql_file = os.path.join(
            backup_root,
            f'db_backup_{timestamp}.sql'
        )

        zip_file = os.path.join(
            backup_root,
            f'db_backup_{timestamp}.zip'
        )

        db = settings.DATABASES['default']

        dump_command = [
            "mysqldump",
            "--no-tablespaces",
            f"-u{db['USER']}",
            f"-p{db['PASSWORD']}",
            db['NAME']
        ]

        # 🔹 Create SQL
        with open(sql_file, "w") as f:
            subprocess.run(dump_command, stdout=f, check=True)

        # 🔹 Zip file
        with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(sql_file, arcname=os.path.basename(sql_file))

        os.remove(sql_file)

        # 🔹 Calculate size in MB
        file_size_mb = round(os.path.getsize(zip_file) / (1024 * 1024), 2)

        # 🔹 Save Backup History
        BackupHistory.objects.create(
            file_name=os.path.basename(zip_file),
            file_path=zip_file,
            file_size_mb=file_size_mb,
            created_by=request.user
        )

        # 🔹 Auto delete backups older than 30 days
        cutoff = timezone.now() - timedelta(days=30)
        old_backups = BackupHistory.objects.filter(created_at__lt=cutoff)

        for backup in old_backups:
            if os.path.exists(backup.file_path):
                os.remove(backup.file_path)
            backup.delete()

        # 🔹 Send Email Notification
        settings_obj, _ = AppSettings.objects.get_or_create(id=1)

        if settings_obj.notification_email:
            subject = "Elite Accounts - Database Backup Completed"

            message = f"""
Hello Admin,

Your database backup was successfully created.

File: {os.path.basename(zip_file)}
Size: {file_size_mb} MB
Date: {now_dt.strftime('%d-%m-%Y')}
Time: {now_dt.strftime('%H:%M')}

Location:
{zip_file}

Elite Accounts System
"""

            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [settings_obj.notification_email],
                fail_silently=False
            )

        # 🔹 Download file
        with open(zip_file, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type='application/zip'
            )
            response['Content-Disposition'] = (
                f'attachment; filename="{os.path.basename(zip_file)}"'
            )
            return response

    except Exception as e:
        messages.error(request, f"Backup failed: {str(e)}")
        return redirect('settings')





import tempfile
import zipfile
import os
import subprocess
from django.views.decorators.http import require_POST
from django.contrib.auth import logout
from django.contrib import messages
from django.shortcuts import redirect
from django.conf import settings
from django.core.mail import send_mail
from .models import AppSettings
from django.utils import timezone
from zoneinfo import ZoneInfo


@login_required
@require_POST
def restore_database(request):
    try:
        uploaded_file = request.FILES.get('backup_file')

        if not uploaded_file:
            messages.error(request, "No file selected.")
            return redirect('settings')

        # 🔒 Logout BEFORE restore
        logout(request)

        restore_time = timezone.now()

        # Save temp zip
        with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as temp_zip:
            for chunk in uploaded_file.chunks():
                temp_zip.write(chunk)
            temp_zip_path = temp_zip.name

        # Extract zip
        with zipfile.ZipFile(temp_zip_path, 'r') as zip_ref:
            zip_ref.extractall("/tmp")
            extracted_files = zip_ref.namelist()

        sql_filename = extracted_files[0]
        sql_path = os.path.join("/tmp", sql_filename)

        db = settings.DATABASES['default']

        restore_command = [
            "mysql",
            f"-u{db['USER']}",
            f"-p{db['PASSWORD']}",
            db['NAME']
        ]

        with open(sql_path, 'r') as f:
            subprocess.run(restore_command, stdin=f, check=True)

        # 🔔 Send Restore Email
        settings_obj, _ = AppSettings.objects.get_or_create(id=1)

        if settings_obj.notification_email:
            send_mail(
                "Elite Accounts - Database Restored",
                f"""
Hello Admin,

Database was successfully restored.

Date: {restore_time.strftime('%d-%m-%Y')}
Time: {restore_time.strftime('%H:%M')}

Elite Accounts System
                """,
                settings.DEFAULT_FROM_EMAIL,
                [settings_obj.notification_email],
                fail_silently=True
            )

        subprocess.run(["systemctl", "restart", "gunicorn-eliteacc"])

        return redirect('/')

    except Exception as e:
        return redirect(f"/settings/?error={str(e)}")



from .models import AppSettings
from django.contrib import messages

def settings_view(request):
    settings_obj, created = AppSettings.objects.get_or_create(id=1)

    if request.method == 'POST':

        # 🔹 Update email
        settings_obj.notification_email = request.POST.get('notification_email')

        # 🔹 Update favicon (if uploaded)
        if request.FILES.get('favicon'):
            settings_obj.favicon = request.FILES.get('favicon')

        settings_obj.save()

        messages.success(request, "Settings updated successfully!")
        return redirect('settings')

    return render(request, 'settings.html', {
        'settings_obj': settings_obj
    })


from django.utils import timezone
from zoneinfo import ZoneInfo
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Sum, Q
from .models import Payment, Expense


from django.utils import timezone
from zoneinfo import ZoneInfo
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Sum, Q

from .models import (
    Payment,
    Expense,
    ActivityLog
)


from zoneinfo import ZoneInfo

from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.db.models import Sum, Q

from .models import (
    Payment,
    Expense,
    ActivityLog
)


# @login_required(login_url='login')
# def activity_view(request):

#     # =====================================================
#     # COMPANY FROM SESSION
#     # =====================================================

#     selected_company_id = request.session.get(
#         'selected_company_id'
#     )

#     if not selected_company_id:
#         return redirect('dashboard')

#     # =====================================================
#     # INDIA TIMEZONE
#     # =====================================================

#     india_tz = ZoneInfo("Asia/Kolkata")

#     current_time_ist = timezone.now().astimezone(india_tz)

#     today = current_time_ist.date()

#     current_month = current_time_ist.month

#     current_year = current_time_ist.year

#     # =====================================================
#     # FILTERS
#     # =====================================================

#     search_query = request.GET.get(
#         'q',
#         ''
#     ).strip()

#     filter_type = request.GET.get(
#         'filter',
#         'today'
#     )

#     log_action = request.GET.get(
#         'log_action',
#         'all'
#     )

#     # =====================================================
#     # PAYMENTS
#     # =====================================================

#     payments = Payment.objects.filter(
#         client__company_id=selected_company_id
#     ).select_related(
#         'client',
#         'bank'
#     )

#     # =====================================================
#     # EXPENSES
#     # =====================================================

#     expenses = Expense.objects.filter(
#         client__company_id=selected_company_id
#     ).select_related(
#         'client',
#         'category',
#         'bank'
#     )

#     # =====================================================
#     # ACTIVITY LOGS
#     # =====================================================

#     activity_logs = ActivityLog.objects.all()

#     # =====================================================
#     # DATE FILTERS
#     # =====================================================

#     # TODAY
#     if filter_type == 'today':

#         payments = payments.filter(
#             payment_date=today
#         )

#         expenses = expenses.filter(
#             expense_date=today
#         )

#         # ✅ FIXED UTC → IST ISSUE
#         activity_logs = activity_logs.filter(
#             created_at__year=current_year,
#             created_at__month=current_month,
#             created_at__day=today.day
#         )

#     # MONTH
#     elif filter_type == 'month':

#         payments = payments.filter(
#             payment_date__month=current_month,
#             payment_date__year=current_year
#         )

#         expenses = expenses.filter(
#             expense_date__month=current_month,
#             expense_date__year=current_year
#         )

#         activity_logs = activity_logs.filter(
#             created_at__month=current_month,
#             created_at__year=current_year
#         )

#     # YEAR
#     elif filter_type == 'year':

#         payments = payments.filter(
#             payment_date__year=current_year
#         )

#         expenses = expenses.filter(
#             expense_date__year=current_year
#         )

#         activity_logs = activity_logs.filter(
#             created_at__year=current_year
#         )

#     # =====================================================
#     # ACTION FILTERS
#     # =====================================================

#     if log_action == 'created':

#         activity_logs = activity_logs.filter(
#             action__icontains='Created'
#         )

#     elif log_action == 'updated':

#         activity_logs = activity_logs.filter(
#             action__icontains='Updated'
#         )

#     elif log_action == 'deleted':

#         activity_logs = activity_logs.filter(
#             action__icontains='Deleted'
#         )

#     # =====================================================
#     # SEARCH FILTER
#     # =====================================================

#     if search_query:

#         payments = payments.filter(

#             Q(client__name__icontains=search_query) |

#             Q(bank__name__icontains=search_query) |

#             Q(amount__icontains=search_query)
#         )

#         expenses = expenses.filter(

#             Q(client__name__icontains=search_query) |

#             Q(category__name__icontains=search_query) |

#             Q(description__icontains=search_query) |

#             Q(amount__icontains=search_query)
#         )

#         activity_logs = activity_logs.filter(

#             Q(action__icontains=search_query) |

#             Q(description__icontains=search_query) |

#             Q(created_by__username__icontains=search_query) |

#             Q(updated_by__username__icontains=search_query) |

#             Q(deleted_by__username__icontains=search_query) |

#             Q(ip_address__icontains=search_query)
#         )

#     # =====================================================
#     # ORDERING
#     # =====================================================

#     payments = payments.order_by(
#         '-payment_date',
#         '-id'
#     )

#     expenses = expenses.order_by(
#         '-expense_date',
#         '-id'
#     )

#     activity_logs = activity_logs.order_by(
#         '-created_at',
#         '-id'
#     )

#     # =====================================================
#     # TOTALS
#     # =====================================================

#     total_income = payments.aggregate(
#         Sum('amount')
#     )['amount__sum'] or 0

#     total_expenses = expenses.aggregate(
#         Sum('amount')
#     )['amount__sum'] or 0

#     balance = total_income - total_expenses

#     # =====================================================
#     # PAGE TITLE
#     # =====================================================

#     if filter_type == 'today':

#         page_title = "Today's Activity"

#     elif filter_type == 'month':

#         page_title = "This Month Activity"

#     elif filter_type == 'year':

#         page_title = "This Year Activity"

#     else:

#         page_title = "Activity"

#     # =====================================================
#     # RENDER
#     # =====================================================

#     return render(
#         request,
#         'activity/today.html',
#         {

#             'payments': payments,

#             'expenses': expenses,

#             'activity_logs': activity_logs,

#             'total_income': total_income,

#             'total_expenses': total_expenses,

#             'balance': balance,

#             'today': today,

#             'current_time_ist': current_time_ist,

#             'search_query': search_query,

#             'filter_type': filter_type,

#             'log_action': log_action,

#             'page_title': page_title,
#         }
#     )




@login_required(login_url='login')
def activity_view(request):

    selected_company_id = request.session.get('selected_company_id')
    if not selected_company_id:
        return redirect('dashboard')

    india_tz = ZoneInfo("Asia/Kolkata")
    current_time_ist = timezone.now().astimezone(india_tz)
    today = current_time_ist.date()
    current_month = current_time_ist.month
    current_year = current_time_ist.year

    # Filters
    search_query = request.GET.get('q', '').strip()
    filter_type = request.GET.get('filter', 'today')
    log_action = request.GET.get('log_action', 'all')

    # Base querysets
    payments = Payment.objects.filter(
        client__company_id=selected_company_id
    ).select_related('client', 'bank')

    expenses = Expense.objects.filter(
        client__company_id=selected_company_id
    ).select_related('client', 'category', 'bank')

    activity_logs = ActivityLog.objects.select_related(
        'created_by', 'updated_by', 'deleted_by'
    ).all()

    # Date filters
    if filter_type == 'today':
        payments = payments.filter(payment_date=today)
        expenses = expenses.filter(expense_date=today)
        activity_logs = activity_logs.filter(
            created_at__date=today
        )
    elif filter_type == 'month':
        payments = payments.filter(
            payment_date__month=current_month,
            payment_date__year=current_year
        )
        expenses = expenses.filter(
            expense_date__month=current_month,
            expense_date__year=current_year
        )
        activity_logs = activity_logs.filter(
            created_at__month=current_month,
            created_at__year=current_year
        )
    elif filter_type == 'year':
        payments = payments.filter(payment_date__year=current_year)
        expenses = expenses.filter(expense_date__year=current_year)
        activity_logs = activity_logs.filter(created_at__year=current_year)

    # Action filter
    if log_action == 'created':
        activity_logs = activity_logs.filter(action__icontains='Created')
    elif log_action == 'updated':
        activity_logs = activity_logs.filter(action__icontains='Updated')
    elif log_action == 'deleted':
        activity_logs = activity_logs.filter(action__icontains='Deleted')

    # Search
    if search_query:
        payments = payments.filter(
            Q(client__name__icontains=search_query) |
            Q(bank__name__icontains=search_query) |
            Q(amount__icontains=search_query)
        )
        expenses = expenses.filter(
            Q(client__name__icontains=search_query) |
            Q(category__name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(amount__icontains=search_query)
        )
        activity_logs = activity_logs.filter(
            Q(action__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(created_by__username__icontains=search_query) |
            Q(ip_address__icontains=search_query)
        )

    # Ordering
    payments = payments.order_by('-payment_date', '-id')
    expenses = expenses.order_by('-expense_date', '-id')
    activity_logs = activity_logs.order_by('-created_at', '-id')

    # Totals
    total_income = payments.aggregate(Sum('amount'))['amount__sum'] or 0
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    balance = total_income - total_expenses

    # Action counts for badges
    all_logs = ActivityLog.objects.all()
    if filter_type == 'today':
        all_logs = all_logs.filter(created_at__date=today)
    elif filter_type == 'month':
        all_logs = all_logs.filter(
            created_at__month=current_month,
            created_at__year=current_year
        )
    elif filter_type == 'year':
        all_logs = all_logs.filter(created_at__year=current_year)

    action_counts = {
        'all': all_logs.count(),
        'created': all_logs.filter(action__icontains='Created').count(),
        'updated': all_logs.filter(action__icontains='Updated').count(),
        'deleted': all_logs.filter(action__icontains='Deleted').count(),
    }

    # Page title
    titles = {
        'today': "Today's Activity",
        'month': "This Month's Activity",
        'year': "This Year's Activity",
    }
    page_title = titles.get(filter_type, "Activity")

    # Current user IP
    x_forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    current_ip = x_forwarded.split(',')[0].strip() if x_forwarded else request.META.get('REMOTE_ADDR', 'Unknown')

    return render(request, 'activity/today.html', {
        'payments': payments,
        'expenses': expenses,
        'activity_logs': activity_logs,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'balance': balance,
        'today': today,
        'current_time_ist': current_time_ist,
        'search_query': search_query,
        'filter_type': filter_type,
        'log_action': log_action,
        'page_title': page_title,
        'action_counts': action_counts,
        'current_ip': current_ip,
    })




@login_required
def backup_history_view(request):
    backups = BackupHistory.objects.order_by('-created_at')

    return render(request, 'backup_history.html', {
        'backups': backups
    })

from django.http import FileResponse, Http404


@login_required
def download_backup(request, backup_id):
    try:
        backup = BackupHistory.objects.get(id=backup_id)

        if not os.path.exists(backup.file_path):
            raise Http404("File not found")

        return FileResponse(
            open(backup.file_path, 'rb'),
            as_attachment=True,
            filename=backup.file_name
        )

    except BackupHistory.DoesNotExist:
        raise Http404("Backup not found")



def help(request):
    return render(request, 'help.html')





#custom error pages


from django.http import HttpResponseBadRequest

from django.conf import settings
import sys
 
def error_404(request, exception):

    error_message = None

    if settings.DEBUG:
        error_message = str(exception)

    return render(request, 'errors/404.html', {
        'error_message': error_message
    }, status=404)


def error_500(request):

    error_message = None

    if settings.DEBUG:
        exc_type, exc_value, exc_traceback = sys.exc_info()
        if exc_value:
            error_message = f"{exc_type.__name__}: {exc_value}"

    return render(request, 'errors/500.html', {
        'error_message': error_message
    }, status=500)



def error_403(request, exception):

    error_message = None

    if settings.DEBUG and exception:
        error_message = f"{type(exception).__name__}: {exception}"

    return render(request, 'errors/403.html', {
        'error_message': error_message
    }, status=403)


def error_400(request, exception):

    error_message = None

    if settings.DEBUG and exception:
        error_message = f"{type(exception).__name__}: {exception}"

    return render(request, 'errors/400.html', {
        'error_message': error_message
    }, status=400)



from django.core.exceptions import PermissionDenied

from django.conf import settings

def error_403(request, exception=None):

    error_message = None

    if settings.DEBUG:
        if exception:
            error_message = f"{type(exception).__name__}: {exception}"
        else:
            error_message = "403 triggered manually (no exception info)"

    return render(request, 'errors/403.html', {
        'error_message': error_message
    }, status=403)


def error_400(request, exception=None):

    error_message = None

    if settings.DEBUG:
        if exception:
            error_message = f"{type(exception).__name__}: {exception}"
        else:
            error_message = "400 Bad Request triggered"

    return render(request, 'errors/400.html', {
        'error_message': error_message
    }, status=400)